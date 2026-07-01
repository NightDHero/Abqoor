import asyncio
import csv
import json
import os
import re
import sqlite3
import tempfile
import threading
from dataclasses import dataclass
from datetime import date, datetime
from pathlib import Path
from typing import Any
from zoneinfo import ZoneInfo, ZoneInfoNotFoundError

import uvicorn
import fitz
from dotenv import load_dotenv
from fastapi import FastAPI, File, Form, Header, HTTPException, UploadFile
from fastapi.responses import FileResponse
from fastapi.staticfiles import StaticFiles
from openpyxl import load_workbook
from telegram import InlineKeyboardButton, InlineKeyboardMarkup, Update
from telegram.error import Forbidden, TelegramError
from telegram.helpers import create_deep_linked_url
from telegram.ext import (
    Application,
    ApplicationBuilder,
    CallbackQueryHandler,
    ChatMemberHandler,
    CommandHandler,
    ContextTypes,
    MessageHandler,
    PollAnswerHandler,
    filters,
)


BASE_DIR = Path(__file__).resolve().parent
STATIC_DIR = BASE_DIR / "static"
VALID_OPTIONS = {"A", "B", "C", "D"}
OPTION_LABELS = {"A": "أ", "B": "ب", "C": "ج", "D": "د"}
POLL_OPTION_ORDER = ["A", "B", "C", "D"]
VALID_CONFLICT_STRATEGIES = {"skip", "replace"}
VALID_IMAGE_SUFFIXES = {".jpg", ".jpeg", ".png", ".webp"}
VALID_ANSWER_SHEET_SUFFIXES = {".xlsx", ".xlsm", ".csv"}
PDF_UPLOAD_CHUNK_SIZE = 1024 * 1024
PDF_RENDER_MAX_EDGE = 1800.0
PDF_RENDER_MAX_PIXELS = 2_500_000.0
SCHEDULER_INTERVAL_SECONDS = 30
MIN_QUESTION_COUNT = 1
MAX_QUESTION_COUNT = 100
CHANNEL_START_PAYLOAD = "from_channel"
MANUAL_REVIEW_SESSION_MARKER = ":mistakebank:"
TODAY_START_LABEL = "ابدأ جلسة اليوم"
TODAY_REDO_LABEL = "أعد جلسة اليوم"
QUESTION_NUMBER_HEADER_ALIASES = {
    "question number",
    "question no",
    "question #",
    "question_number",
    "question_no",
    "q number",
    "q no",
    "q#",
    "qnum",
    "question id",
    "رقم السؤال",
    "رقم",
}
ANSWER_HEADER_ALIASES = {
    "answer",
    "answers",
    "correct answer",
    "correct option",
    "correct_option",
    "answer key",
    "ans",
    "الإجابة",
    "الاجابة",
    "الإجابة الصحيحة",
    "الاجابة الصحيحة",
}

ARABIC_WEEKDAY_ORDER: list[tuple[int, str]] = [
    (5, "السبت"),
    (6, "الأحد"),
    (0, "الاثنين"),
    (1, "الثلاثاء"),
    (2, "الأربعاء"),
    (3, "الخميس"),
    (4, "الجمعة"),
]
ARABIC_WEEKDAY_LABELS = {weekday: label for weekday, label in ARABIC_WEEKDAY_ORDER}
ARABIC_DIGITS_TRANSLATION = str.maketrans("٠١٢٣٤٥٦٧٨٩۰۱۲۳۴۵۶۷۸۹", "01234567890123456789")


@dataclass(slots=True)
class Settings:
    telegram_bot_token: str
    admin_password: str
    host: str = "127.0.0.1"
    port: int = 8000
    storage_dir: Path = BASE_DIR
    timezone_name: str = "Asia/Riyadh"
    telegram_enabled: bool = True

    @property
    def data_dir(self) -> Path:
        return self.storage_dir / "data"

    @property
    def uploads_dir(self) -> Path:
        return self.storage_dir / "uploads"

    @property
    def database_path(self) -> Path:
        return self.data_dir / "study.db"

    @property
    def legacy_question_file(self) -> Path:
        return self.data_dir / "questions.json"

    @property
    def timezone(self) -> ZoneInfo:
        return ZoneInfo(self.timezone_name)


class StudyRepository:
    def __init__(self, database_path: Path, uploads_dir: Path, legacy_question_file: Path) -> None:
        self.database_path = database_path
        self.uploads_dir = uploads_dir
        self.legacy_question_file = legacy_question_file
        self._lock = threading.Lock()
        self._connection: sqlite3.Connection | None = None

    def initialize(self) -> None:
        self.database_path.parent.mkdir(parents=True, exist_ok=True)
        self.uploads_dir.mkdir(parents=True, exist_ok=True)
        self._connection = sqlite3.connect(self.database_path, check_same_thread=False)
        self._connection.row_factory = sqlite3.Row
        self._connection.execute("PRAGMA foreign_keys = ON")
        self._create_tables()
        self._apply_migrations()
        self._migrate_legacy_questions_if_needed()

    def close(self) -> None:
        if self._connection is not None:
            self._connection.close()
            self._connection = None

    @property
    def connection(self) -> sqlite3.Connection:
        if self._connection is None:
            raise RuntimeError("Repository has not been initialized.")
        return self._connection

    def _create_tables(self) -> None:
        with self._lock, self.connection:
            self.connection.executescript(
                """
                CREATE TABLE IF NOT EXISTS users (
                    id INTEGER PRIMARY KEY AUTOINCREMENT,
                    telegram_user_id INTEGER NOT NULL UNIQUE,
                    username TEXT,
                    first_name TEXT,
                    last_name TEXT,
                    display_name TEXT NOT NULL,
                    language_code TEXT,
                    created_at TEXT NOT NULL,
                    updated_at TEXT NOT NULL
                );

                CREATE TABLE IF NOT EXISTS study_plans (
                    id INTEGER PRIMARY KEY AUTOINCREMENT,
                    user_id INTEGER NOT NULL UNIQUE REFERENCES users(id) ON DELETE CASCADE,
                    reminder_time TEXT NOT NULL,
                    question_count INTEGER NOT NULL,
                    review_weekday INTEGER,
                    timezone_name TEXT NOT NULL,
                    is_active INTEGER NOT NULL DEFAULT 1,
                    onboarding_completed INTEGER NOT NULL DEFAULT 0,
                    created_at TEXT NOT NULL,
                    updated_at TEXT NOT NULL
                );

                CREATE TABLE IF NOT EXISTS plan_days (
                    plan_id INTEGER NOT NULL REFERENCES study_plans(id) ON DELETE CASCADE,
                    weekday INTEGER NOT NULL,
                    PRIMARY KEY(plan_id, weekday)
                );

                CREATE TABLE IF NOT EXISTS questions (
                    id INTEGER PRIMARY KEY AUTOINCREMENT,
                    question_number INTEGER,
                    caption TEXT NOT NULL DEFAULT '',
                    topic TEXT NOT NULL DEFAULT 'عام',
                    difficulty TEXT NOT NULL DEFAULT 'متوسط',
                    correct_option TEXT NOT NULL,
                    image_path TEXT NOT NULL,
                    original_filename TEXT NOT NULL,
                    is_active INTEGER NOT NULL DEFAULT 1,
                    created_at TEXT NOT NULL
                );

                CREATE TABLE IF NOT EXISTS study_sessions (
                    id INTEGER PRIMARY KEY AUTOINCREMENT,
                    user_id INTEGER NOT NULL REFERENCES users(id) ON DELETE CASCADE,
                    scheduled_for TEXT NOT NULL,
                    session_kind TEXT NOT NULL DEFAULT 'study',
                    status TEXT NOT NULL DEFAULT 'pending',
                    question_count INTEGER NOT NULL,
                    current_position INTEGER NOT NULL DEFAULT 1,
                    notice_message_id INTEGER,
                    view_message_id INTEGER,
                    active_poll_id TEXT,
                    active_poll_message_id INTEGER,
                    active_poll_question_id INTEGER,
                    created_at TEXT NOT NULL,
                    delivered_at TEXT,
                    started_at TEXT,
                    completed_at TEXT,
                    UNIQUE(user_id, scheduled_for)
                );

                CREATE TABLE IF NOT EXISTS session_questions (
                    id INTEGER PRIMARY KEY AUTOINCREMENT,
                    session_id INTEGER NOT NULL REFERENCES study_sessions(id) ON DELETE CASCADE,
                    question_id INTEGER NOT NULL REFERENCES questions(id) ON DELETE RESTRICT,
                    position INTEGER NOT NULL,
                    correct_option TEXT NOT NULL,
                    selected_option TEXT,
                    is_correct INTEGER,
                    answered_at TEXT,
                    delivery_message_id INTEGER,
                    UNIQUE(session_id, position),
                    UNIQUE(session_id, question_id)
                );

                CREATE TABLE IF NOT EXISTS mistake_bank (
                    user_id INTEGER NOT NULL REFERENCES users(id) ON DELETE CASCADE,
                    question_id INTEGER NOT NULL REFERENCES questions(id) ON DELETE RESTRICT,
                    first_wrong_at TEXT NOT NULL,
                    last_wrong_at TEXT NOT NULL,
                    wrong_count INTEGER NOT NULL DEFAULT 1,
                    correct_count INTEGER NOT NULL DEFAULT 0,
                    last_correct_at TEXT,
                    is_active INTEGER NOT NULL DEFAULT 1,
                    PRIMARY KEY(user_id, question_id)
                );
                """
            )

    def _apply_migrations(self) -> None:
        self._ensure_column("questions", "question_number INTEGER")
        self._ensure_column("study_plans", "review_weekday INTEGER")
        self._ensure_column("study_sessions", "session_kind TEXT NOT NULL DEFAULT 'study'")
        self._ensure_column("study_sessions", "current_position INTEGER NOT NULL DEFAULT 1")
        self._ensure_column("study_sessions", "notice_message_id INTEGER")
        self._ensure_column("study_sessions", "view_message_id INTEGER")
        self._ensure_column("study_sessions", "active_poll_id TEXT")
        self._ensure_column("study_sessions", "active_poll_message_id INTEGER")
        self._ensure_column("study_sessions", "active_poll_question_id INTEGER")

        with self._lock, self.connection:
            self.connection.execute(
                "UPDATE questions SET question_number = id WHERE question_number IS NULL OR question_number < 1"
            )
            self.connection.execute(
                """
                CREATE UNIQUE INDEX IF NOT EXISTS idx_questions_active_question_number
                ON questions(question_number)
                WHERE is_active = 1 AND question_number IS NOT NULL
                """
            )
            self.connection.execute(
                "UPDATE study_sessions SET session_kind = 'study' WHERE session_kind IS NULL OR TRIM(session_kind) = ''"
            )
            self.connection.execute(
                "UPDATE study_sessions SET current_position = 1 WHERE current_position IS NULL OR current_position < 1"
            )

    def _ensure_column(self, table_name: str, column_definition: str) -> None:
        column_name = column_definition.split()[0]
        existing_columns = {
            str(row["name"])
            for row in self.connection.execute(f"PRAGMA table_info({table_name})").fetchall()
        }
        if column_name in existing_columns:
            return

        with self._lock, self.connection:
            self.connection.execute(f"ALTER TABLE {table_name} ADD COLUMN {column_definition}")

    def _migrate_legacy_questions_if_needed(self) -> None:
        if not self.legacy_question_file.exists():
            return

        question_count = self.connection.execute("SELECT COUNT(*) FROM questions").fetchone()[0]
        if question_count:
            return

        raw_payload = json.loads(self.legacy_question_file.read_text(encoding="utf-8-sig"))
        items = raw_payload.get("questions", [])
        if not isinstance(items, list):
            return

        used_question_numbers: set[int] = set()

        with self._lock, self.connection:
            for index, item in enumerate(items, start=1):
                image_path = str(item.get("image_path", "")).strip()
                image_name = Path(image_path).name
                fallback_path = self.uploads_dir / image_name if image_name else self.uploads_dir / ""
                resolved_image_path = image_path
                if image_name and not Path(image_path).exists() and fallback_path.exists():
                    resolved_image_path = str(fallback_path)

                if not resolved_image_path:
                    continue

                question_number = parse_question_number_cell(item.get("question_number"))
                if question_number is None:
                    question_number = parse_question_number_cell(item.get("id"))
                if question_number is None:
                    question_number = index
                while question_number in used_question_numbers:
                    question_number += 1
                used_question_numbers.add(question_number)

                self.connection.execute(
                    """
                    INSERT INTO questions (
                        question_number,
                        caption,
                        topic,
                        difficulty,
                        correct_option,
                        image_path,
                        original_filename,
                        is_active,
                        created_at
                    ) VALUES (?, ?, ?, ?, ?, ?, ?, 1, ?)
                    """,
                    (
                        question_number,
                        str(item.get("caption", "")).strip(),
                        "مستورد",
                        "متوسط",
                        str(item.get("correct_option", "A")).strip().upper(),
                        resolved_image_path,
                        str(item.get("original_filename", image_name or "legacy-upload")),
                        str(item.get("created_at", utc_now_iso())),
                    ),
                )

    def upsert_telegram_user(self, telegram_user: Any) -> int:
        display_name = " ".join(
            part for part in [telegram_user.first_name, telegram_user.last_name] if part
        ).strip() or telegram_user.username or "طالب"
        timestamp = utc_now_iso()

        with self._lock, self.connection:
            existing = self.connection.execute(
                "SELECT id FROM users WHERE telegram_user_id = ?",
                (telegram_user.id,),
            ).fetchone()

            if existing is None:
                cursor = self.connection.execute(
                    """
                    INSERT INTO users (
                        telegram_user_id,
                        username,
                        first_name,
                        last_name,
                        display_name,
                        language_code,
                        created_at,
                        updated_at
                    ) VALUES (?, ?, ?, ?, ?, ?, ?, ?)
                    """,
                    (
                        telegram_user.id,
                        telegram_user.username,
                        telegram_user.first_name,
                        telegram_user.last_name,
                        display_name,
                        telegram_user.language_code,
                        timestamp,
                        timestamp,
                    ),
                )
                return int(cursor.lastrowid)

            self.connection.execute(
                """
                UPDATE users
                SET username = ?, first_name = ?, last_name = ?, display_name = ?, language_code = ?, updated_at = ?
                WHERE telegram_user_id = ?
                """,
                (
                    telegram_user.username,
                    telegram_user.first_name,
                    telegram_user.last_name,
                    display_name,
                    telegram_user.language_code,
                    timestamp,
                    telegram_user.id,
                ),
            )
            return int(existing["id"])

    def save_plan(
        self,
        telegram_user: Any,
        *,
        weekdays: list[int],
        reminder_time: str,
        question_count: int,
        review_weekday: int | None,
        timezone_name: str,
    ) -> dict[str, Any]:
        user_id = self.upsert_telegram_user(telegram_user)
        timestamp = utc_now_iso()

        with self._lock, self.connection:
            existing_plan = self.connection.execute(
                "SELECT id FROM study_plans WHERE user_id = ?",
                (user_id,),
            ).fetchone()

            if existing_plan is None:
                cursor = self.connection.execute(
                    """
                    INSERT INTO study_plans (
                        user_id,
                        reminder_time,
                        question_count,
                        review_weekday,
                        timezone_name,
                        is_active,
                        onboarding_completed,
                        created_at,
                        updated_at
                    ) VALUES (?, ?, ?, ?, ?, 1, 1, ?, ?)
                    """,
                    (
                        user_id,
                        reminder_time,
                        question_count,
                        review_weekday,
                        timezone_name,
                        timestamp,
                        timestamp,
                    ),
                )
                plan_id = int(cursor.lastrowid)
            else:
                plan_id = int(existing_plan["id"])
                self.connection.execute(
                    """
                    UPDATE study_plans
                    SET reminder_time = ?, question_count = ?, review_weekday = ?, timezone_name = ?, is_active = 1,
                        onboarding_completed = 1, updated_at = ?
                    WHERE id = ?
                    """,
                    (reminder_time, question_count, review_weekday, timezone_name, timestamp, plan_id),
                )
                self.connection.execute("DELETE FROM plan_days WHERE plan_id = ?", (plan_id,))

            self.connection.executemany(
                "INSERT INTO plan_days (plan_id, weekday) VALUES (?, ?)",
                [(plan_id, weekday) for weekday in sorted(set(weekdays))],
            )

        plan = self.get_plan_by_telegram_id(telegram_user.id) or {}
        if plan:
            self.sync_today_session_to_plan(
                plan,
                datetime.now(ZoneInfo(str(timezone_name))).date(),
            )
        return plan

    def set_plan_active(self, telegram_user_id: int, is_active: bool) -> dict[str, Any] | None:
        plan = self.get_plan_by_telegram_id(telegram_user_id)
        if plan is None:
            return None

        with self._lock, self.connection:
            self.connection.execute(
                "UPDATE study_plans SET is_active = ?, updated_at = ? WHERE id = ?",
                (1 if is_active else 0, utc_now_iso(), plan["plan_id"]),
            )

        return self.get_plan_by_telegram_id(telegram_user_id)

    def reset_user(self, telegram_user_id: int) -> bool:
        row = self.connection.execute(
            "SELECT id FROM users WHERE telegram_user_id = ?",
            (telegram_user_id,),
        ).fetchone()
        if row is None:
            return False

        with self._lock, self.connection:
            self.connection.execute(
                "DELETE FROM users WHERE telegram_user_id = ?",
                (telegram_user_id,),
            )

        return True

    def get_plan_by_telegram_id(self, telegram_user_id: int) -> dict[str, Any] | None:
        row = self.connection.execute(
            """
            SELECT
                sp.id AS plan_id,
                sp.user_id,
                sp.reminder_time,
                sp.question_count,
                sp.review_weekday,
                sp.timezone_name,
                sp.is_active,
                sp.onboarding_completed,
                u.telegram_user_id,
                u.display_name,
                GROUP_CONCAT(pd.weekday) AS weekdays
            FROM study_plans sp
            JOIN users u ON u.id = sp.user_id
            LEFT JOIN plan_days pd ON pd.plan_id = sp.id
            WHERE u.telegram_user_id = ?
            GROUP BY sp.id
            """,
            (telegram_user_id,),
        ).fetchone()
        if row is None:
            return None
        return self._plan_row_to_dict(row)

    def get_dashboard_stats(self) -> dict[str, int]:
        user_count = int(self.connection.execute("SELECT COUNT(*) FROM users").fetchone()[0])
        active_plans = int(
            self.connection.execute(
                "SELECT COUNT(*) FROM study_plans WHERE is_active = 1 AND onboarding_completed = 1"
            ).fetchone()[0]
        )
        question_count = int(
            self.connection.execute(
                "SELECT COUNT(*) FROM questions WHERE is_active = 1 AND correct_option IN ('A', 'B', 'C', 'D')"
            ).fetchone()[0]
        )
        return {
            "user_count": user_count,
            "active_plans": active_plans,
            "question_count": question_count,
        }

    def _get_question_row_by_number(self, question_number: int, *, active_only: bool = True) -> sqlite3.Row | None:
        status_clause = "AND q.is_active = 1" if active_only else ""
        row = self.connection.execute(
            f"""
            SELECT q.*, COUNT(sq.id) AS assignment_count
            FROM questions q
            LEFT JOIN session_questions sq ON sq.question_id = q.id
            WHERE q.question_number = ? {status_clause}
            GROUP BY q.id
            ORDER BY q.id DESC
            LIMIT 1
            """,
            (question_number,),
        ).fetchone()
        return row

    def get_question_by_number_public(self, question_number: int, *, active_only: bool = True) -> dict[str, Any] | None:
        row = self._get_question_row_by_number(question_number, active_only=active_only)
        if row is None:
            return None
        return self._question_row_to_public(row)

    def list_active_questions_by_numbers(self, question_numbers: list[int]) -> list[dict[str, Any]]:
        normalized_numbers = sorted({int(number) for number in question_numbers if int(number) > 0})
        if not normalized_numbers:
            return []

        placeholders = ",".join("?" for _ in normalized_numbers)
        rows = self.connection.execute(
            f"""
            SELECT q.*, COUNT(sq.id) AS assignment_count
            FROM questions q
            LEFT JOIN session_questions sq ON sq.question_id = q.id
            WHERE q.is_active = 1 AND q.question_number IN ({placeholders})
            GROUP BY q.id
            ORDER BY q.question_number ASC, q.id DESC
            """,
            tuple(normalized_numbers),
        ).fetchall()
        return [self._question_row_to_public(row) for row in rows]

    def archive_questions(self, question_ids: list[int]) -> dict[str, Any]:
        normalized_ids = [int(question_id) for question_id in dict.fromkeys(question_ids) if int(question_id) > 0]
        if not normalized_ids:
            return {
                "ok": True,
                "archived_question_ids": [],
                "archived_question_numbers": [],
                "archived_count": 0,
            }

        placeholders = ",".join("?" for _ in normalized_ids)
        rows = self.connection.execute(
            f"""
            SELECT id, question_number
            FROM questions
            WHERE is_active = 1 AND id IN ({placeholders})
            ORDER BY question_number ASC, id ASC
            """,
            tuple(normalized_ids),
        ).fetchall()
        if not rows:
            return {
                "ok": True,
                "archived_question_ids": [],
                "archived_question_numbers": [],
                "archived_count": 0,
            }

        archived_ids = [int(row["id"]) for row in rows]
        archived_numbers = [int(row["question_number"] or row["id"]) for row in rows]
        archive_placeholders = ",".join("?" for _ in archived_ids)

        with self._lock, self.connection:
            self.connection.execute(
                f"UPDATE questions SET is_active = 0 WHERE id IN ({archive_placeholders})",
                tuple(archived_ids),
            )

        return {
            "ok": True,
            "archived_question_ids": archived_ids,
            "archived_question_numbers": archived_numbers,
            "archived_count": len(archived_ids),
        }

    def archive_all_questions(self) -> dict[str, Any]:
        rows = self.connection.execute(
            """
            SELECT id, question_number
            FROM questions
            WHERE is_active = 1
            ORDER BY question_number ASC, id ASC
            """
        ).fetchall()
        if not rows:
            return {
                "ok": True,
                "archived_question_ids": [],
                "archived_question_numbers": [],
                "archived_count": 0,
            }

        archived_ids = [int(row["id"]) for row in rows]
        archived_numbers = [int(row["question_number"] or row["id"]) for row in rows]

        with self._lock, self.connection:
            self.connection.execute("UPDATE questions SET is_active = 0 WHERE is_active = 1")

        return {
            "ok": True,
            "archived_question_ids": archived_ids,
            "archived_question_numbers": archived_numbers,
            "archived_count": len(archived_ids),
        }

    def restore_questions(self, question_ids: list[int]) -> dict[str, Any]:
        normalized_ids = [int(question_id) for question_id in dict.fromkeys(question_ids) if int(question_id) > 0]
        if not normalized_ids:
            return {
                "ok": True,
                "restored_question_ids": [],
                "restored_question_numbers": [],
                "restored_count": 0,
            }

        placeholders = ",".join("?" for _ in normalized_ids)
        rows = self.connection.execute(
            f"""
            SELECT id, question_number
            FROM questions
            WHERE is_active = 0 AND id IN ({placeholders})
            ORDER BY question_number ASC, id ASC
            """,
            tuple(normalized_ids),
        ).fetchall()
        if not rows:
            return {
                "ok": True,
                "restored_question_ids": [],
                "restored_question_numbers": [],
                "restored_count": 0,
            }

        restored_ids = [int(row["id"]) for row in rows]
        restored_numbers = [int(row["question_number"] or row["id"]) for row in rows]
        conflict_rows = self.list_active_questions_by_numbers(restored_numbers)
        if conflict_rows:
            conflict_numbers = [str(int(item["question_number"])) for item in conflict_rows]
            raise HTTPException(
                status_code=409,
                detail=f"لا يمكن التراجع الآن لأن أرقام الأسئلة التالية مستخدمة حاليًا: {'، '.join(conflict_numbers)}.",
            )

        restore_placeholders = ",".join("?" for _ in restored_ids)
        with self._lock, self.connection:
            self.connection.execute(
                f"UPDATE questions SET is_active = 1 WHERE id IN ({restore_placeholders})",
                tuple(restored_ids),
            )

        return {
            "ok": True,
            "restored_question_ids": restored_ids,
            "restored_question_numbers": restored_numbers,
            "restored_count": len(restored_ids),
        }

    def save_question(
        self,
        *,
        question_number: int,
        caption: str,
        topic: str,
        difficulty: str,
        correct_option: str,
        original_filename: str,
        image_bytes: bytes,
    ) -> dict[str, Any]:
        normalized_question_number = int(question_number)
        if normalized_question_number <= 0:
            raise HTTPException(status_code=422, detail="رقم السؤال يجب أن يكون رقمًا صحيحًا أكبر من صفر.")
        normalized_topic = topic.strip() or "عام"
        normalized_difficulty = difficulty.strip() or "متوسط"
        normalized_option = correct_option.strip().upper()
        if normalized_option and normalized_option not in VALID_OPTIONS:
            raise HTTPException(status_code=422, detail="الإجابة الصحيحة يجب أن تكون أ أو ب أو ج أو د.")
        timestamp = utc_now_iso()
        suffix = self._resolve_suffix(original_filename)

        existing_question = self._get_question_row_by_number(normalized_question_number)
        if existing_question is not None:
            raise HTTPException(status_code=409, detail=f"رقم السؤال {normalized_question_number} موجود بالفعل.")

        with self._lock, self.connection:
            cursor = self.connection.execute(
                """
                INSERT INTO questions (
                    question_number,
                    caption,
                    topic,
                    difficulty,
                    correct_option,
                    image_path,
                    original_filename,
                    is_active,
                    created_at
                ) VALUES (?, ?, ?, ?, ?, ?, ?, 1, ?)
                """,
                (
                    normalized_question_number,
                    caption.strip(),
                    normalized_topic,
                    normalized_difficulty,
                    normalized_option,
                    "",
                    original_filename,
                    timestamp,
                ),
            )
            question_id = int(cursor.lastrowid)
            image_path = self.uploads_dir / f"question_{question_id}{suffix}"
            image_path.write_bytes(image_bytes)
            self.connection.execute(
                "UPDATE questions SET image_path = ? WHERE id = ?",
                (str(image_path), question_id),
            )

        return self.get_question_public(question_id)

    def _render_pdf_page_to_png(self, page: fitz.Page) -> bytes:
        page_width = max(float(page.rect.width), 1.0)
        page_height = max(float(page.rect.height), 1.0)
        edge_limit_scale = PDF_RENDER_MAX_EDGE / max(page_width, page_height)
        pixel_limit_scale = (PDF_RENDER_MAX_PIXELS / (page_width * page_height)) ** 0.5
        render_scales: list[float] = []

        for requested_scale in (2.0, 1.5, 1.0, 0.8, 0.6):
            safe_scale = min(requested_scale, edge_limit_scale, pixel_limit_scale)
            if safe_scale <= 0:
                continue
            rounded_scale = round(safe_scale, 3)
            if rounded_scale not in render_scales:
                render_scales.append(rounded_scale)

        if not render_scales:
            render_scales.append(0.5)

        last_error: Exception | None = None

        for scale in render_scales:
            try:
                pixmap = page.get_pixmap(
                    matrix=fitz.Matrix(scale, scale),
                    colorspace=fitz.csRGB,
                    alpha=False,
                    annots=False,
                )
                page_bytes = pixmap.tobytes("png")
                pixmap = None
                return page_bytes
            except Exception as exc:
                last_error = exc

        if last_error is None:
            raise RuntimeError("تعذر تحويل الصفحة إلى صورة.")
        raise last_error

    def import_pdf_questions(
        self,
        *,
        original_filename: str,
        pdf_path: Path,
        starting_question_number: int,
        topic: str = "عام",
        difficulty: str = "متوسط",
        answer_key_by_question_number: dict[int, str] | None = None,
        existing_question_strategy: str = "error",
    ) -> tuple[list[dict[str, Any]], list[int], list[int], int, list[int], list[int]]:
        try:
            document = fitz.open(str(pdf_path))
        except Exception as exc:
            raise HTTPException(status_code=400, detail="تعذر قراءة ملف PDF المرفوع.") from exc

        imported_questions: list[dict[str, Any]] = []
        failed_pages: list[int] = []
        unanswered_question_numbers: list[int] = []
        skipped_existing_question_numbers: list[int] = []
        replaced_question_numbers: list[int] = []
        ready_count = 0
        pdf_stem = Path(original_filename or "imported").stem or "imported"

        try:
            if document.needs_pass and not document.authenticate(""):
                raise HTTPException(status_code=400, detail="ملف PDF محمي بكلمة مرور ولا يمكن تحويله إلى صور.")

            for page_index in range(document.page_count):
                question_number = starting_question_number + page_index
                existing_question = self._get_question_row_by_number(question_number)
                if existing_question is not None:
                    if existing_question_strategy == "skip":
                        skipped_existing_question_numbers.append(question_number)
                        continue
                    if existing_question_strategy == "replace":
                        with self._lock, self.connection:
                            self.connection.execute(
                                "UPDATE questions SET is_active = 0 WHERE id = ?",
                                (int(existing_question["id"]),),
                            )
                        replaced_question_numbers.append(question_number)
                    else:
                        raise HTTPException(status_code=409, detail=f"رقم السؤال {question_number} موجود بالفعل.")

                page = document.load_page(page_index)
                try:
                    page_bytes = self._render_pdf_page_to_png(page)
                except Exception:
                    failed_pages.append(page_index + 1)
                    continue

                correct_option = ""
                if answer_key_by_question_number is not None:
                    correct_option = answer_key_by_question_number.get(question_number, "")
                    if not correct_option:
                        unanswered_question_numbers.append(question_number)
                    else:
                        ready_count += 1

                imported_questions.append(
                    self.save_question(
                        question_number=question_number,
                        caption="",
                        topic=topic,
                        difficulty=difficulty,
                        correct_option=correct_option,
                        original_filename=f"{pdf_stem}_page_{page_index + 1}.png",
                        image_bytes=page_bytes,
                    )
                )
                if hasattr(fitz, "TOOLS"):
                    try:
                        fitz.TOOLS.store_shrink(100)
                    except Exception:
                        pass
        finally:
            document.close()

        if not imported_questions:
            if skipped_existing_question_numbers:
                return (
                    imported_questions,
                    failed_pages,
                    unanswered_question_numbers,
                    ready_count,
                    skipped_existing_question_numbers,
                    replaced_question_numbers,
                )
            raise HTTPException(
                status_code=400,
                detail="تعذر تحويل صفحات ملف PDF إلى صور. جرّب حفظه من جديد أو تصديره بدون حماية.",
            )

        return (
            imported_questions,
            failed_pages,
            unanswered_question_numbers,
            ready_count,
            skipped_existing_question_numbers,
            replaced_question_numbers,
        )

    def update_question_correct_option(self, question_id: int, correct_option: str) -> dict[str, Any] | None:
        normalized_option = correct_option.strip().upper()
        if normalized_option not in VALID_OPTIONS:
            raise HTTPException(status_code=422, detail="الإجابة الصحيحة يجب أن تكون أ أو ب أو ج أو د.")

        row = self.connection.execute(
            "SELECT id FROM questions WHERE id = ? AND is_active = 1",
            (question_id,),
        ).fetchone()
        if row is None:
            return None

        with self._lock, self.connection:
            self.connection.execute(
                "UPDATE questions SET correct_option = ? WHERE id = ?",
                (normalized_option, question_id),
            )

        return self.get_question_public(question_id)

    def list_questions(self, limit: int = 48) -> list[dict[str, Any]]:
        rows = self.connection.execute(
            """
            SELECT q.*, COUNT(sq.id) AS assignment_count
            FROM questions q
            LEFT JOIN session_questions sq ON sq.question_id = q.id
            WHERE q.is_active = 1
            GROUP BY q.id
            ORDER BY q.question_number DESC, q.id DESC
            LIMIT ?
            """,
            (limit,),
        ).fetchall()
        return [self._question_row_to_public(row) for row in rows]

    def get_question_public(self, question_id: int) -> dict[str, Any]:
        row = self.connection.execute(
            """
            SELECT q.*, COUNT(sq.id) AS assignment_count
            FROM questions q
            LEFT JOIN session_questions sq ON sq.question_id = q.id
            WHERE q.id = ?
            GROUP BY q.id
            """,
            (question_id,),
        ).fetchone()
        if row is None:
            raise HTTPException(status_code=404, detail="السؤال غير موجود.")
        return self._question_row_to_public(row)

    def archive_question(self, question_id: int) -> dict[str, Any] | None:
        archived = self.archive_questions([question_id])
        if not archived["archived_question_ids"]:
            return None
        return archived

    def list_due_plans(self, now_local: datetime) -> list[dict[str, Any]]:
        current_date = now_local.date().isoformat()
        current_time = now_local.strftime("%H:%M")
        current_weekday = now_local.weekday()
        rows = self.connection.execute(
            """
            SELECT
                sp.id AS plan_id,
                sp.user_id,
                sp.reminder_time,
                sp.question_count,
                sp.review_weekday,
                sp.timezone_name,
                sp.is_active,
                sp.onboarding_completed,
                u.telegram_user_id,
                u.display_name,
                GROUP_CONCAT(pd.weekday) AS weekdays
            FROM study_plans sp
            JOIN users u ON u.id = sp.user_id
            LEFT JOIN plan_days pd ON pd.plan_id = sp.id
            WHERE sp.is_active = 1 AND sp.onboarding_completed = 1
            GROUP BY sp.id
            """
        ).fetchall()

        due_plans: list[dict[str, Any]] = []
        for row in rows:
            plan = self._plan_row_to_dict(row)
            if current_weekday not in plan["weekdays"]:
                continue
            if plan["reminder_time"] > current_time:
                continue

            existing = self.connection.execute(
                "SELECT id FROM study_sessions WHERE user_id = ? AND scheduled_for = ?",
                (plan["user_id"], current_date),
            ).fetchone()
            if existing is not None:
                continue

            due_plans.append(plan)

        return due_plans

    def create_session_for_user(self, telegram_user_id: int, scheduled_for: date) -> dict[str, Any] | None:
        plan = self.get_plan_by_telegram_id(telegram_user_id)
        if plan is None or not plan["onboarding_completed"]:
            return None

        scheduled_for_text = scheduled_for.isoformat()
        self._expire_old_sessions(plan["user_id"], scheduled_for_text)

        existing = self.connection.execute(
            """
            SELECT ss.*, u.telegram_user_id, u.display_name
            FROM study_sessions ss
            JOIN users u ON u.id = ss.user_id
            WHERE ss.user_id = ? AND ss.scheduled_for = ?
            """,
            (plan["user_id"], scheduled_for_text),
        ).fetchone()
        if existing is not None:
            return self._session_row_to_dict(existing)

        session_kind, selected_questions = self._select_questions_for_plan(plan, scheduled_for)

        if not selected_questions:
            return None

        return self._create_session_record(plan["user_id"], scheduled_for_text, session_kind, selected_questions)

    def sync_today_session_to_plan(self, plan: dict[str, Any], scheduled_for: date) -> dict[str, Any] | None:
        scheduled_for_text = scheduled_for.isoformat()
        row = self.connection.execute(
            """
            SELECT ss.*, u.telegram_user_id, u.display_name
            FROM study_sessions ss
            JOIN users u ON u.id = ss.user_id
            WHERE ss.user_id = ? AND ss.scheduled_for = ?
            """,
            (plan["user_id"], scheduled_for_text),
        ).fetchone()
        if row is None:
            return None

        session = self._session_row_to_dict(row)
        desired_count = max(1, int(plan["question_count"]))
        current_count = int(session["question_count"])
        if desired_count <= current_count:
            return session

        existing_question_ids = self._get_session_question_ids(int(session["session_id"]))
        additional_count = desired_count - current_count
        if session["session_kind"] == "review":
            additional_questions = self._select_mistake_questions_for_user(
                int(plan["user_id"]),
                additional_count,
                exclude_question_ids=existing_question_ids,
            )
        else:
            additional_questions = self._select_questions_for_user(
                int(plan["user_id"]),
                additional_count,
                exclude_question_ids=existing_question_ids,
            )

        if not additional_questions:
            return session

        first_new_position = current_count + 1
        new_total = current_count + len(additional_questions)
        with self._lock, self.connection:
            self.connection.executemany(
                """
                INSERT INTO session_questions (
                    session_id,
                    question_id,
                    position,
                    correct_option
                ) VALUES (?, ?, ?, ?)
                """,
                [
                    (
                        int(session["session_id"]),
                        int(item["id"]),
                        position,
                        str(item["correct_option"]),
                    )
                    for position, item in enumerate(additional_questions, start=first_new_position)
                ],
            )
            if session["status"] == "completed":
                self.connection.execute(
                    """
                    UPDATE study_sessions
                    SET question_count = ?,
                        current_position = ?,
                        status = 'pending',
                        started_at = NULL,
                        completed_at = NULL
                    WHERE id = ?
                    """,
                    (new_total, first_new_position, int(session["session_id"])),
                )
            else:
                self.connection.execute(
                    "UPDATE study_sessions SET question_count = ? WHERE id = ?",
                    (new_total, int(session["session_id"])),
                )

        return self.get_session_by_id(int(session["session_id"]))

    def reset_completed_session_for_date(
        self,
        telegram_user_id: int,
        scheduled_for: date,
    ) -> dict[str, Any] | None:
        plan = self.get_plan_by_telegram_id(telegram_user_id)
        if plan is None or not plan["onboarding_completed"]:
            return None

        session = self.get_session_for_user_and_date(telegram_user_id, scheduled_for)
        if session is None or session["status"] != "completed":
            return None

        session_kind, selected_questions = self._select_questions_for_plan(plan, scheduled_for)
        if not selected_questions:
            return None

        with self._lock, self.connection:
            self.connection.execute(
                "DELETE FROM session_questions WHERE session_id = ?",
                (int(session["session_id"]),),
            )
            self.connection.execute(
                """
                UPDATE study_sessions
                SET session_kind = ?,
                    status = 'pending',
                    question_count = ?,
                    current_position = 1,
                    notice_message_id = NULL,
                    view_message_id = NULL,
                    active_poll_id = NULL,
                    active_poll_message_id = NULL,
                    active_poll_question_id = NULL,
                    delivered_at = NULL,
                    started_at = NULL,
                    completed_at = NULL
                WHERE id = ?
                """,
                (session_kind, len(selected_questions), int(session["session_id"])),
            )
            self.connection.executemany(
                """
                INSERT INTO session_questions (
                    session_id,
                    question_id,
                    position,
                    correct_option
                ) VALUES (?, ?, ?, ?)
                """,
                [
                    (
                        int(session["session_id"]),
                        int(item["id"]),
                        position,
                        str(item["correct_option"]),
                    )
                    for position, item in enumerate(selected_questions, start=1)
                ],
            )

        return self.get_session_by_id(int(session["session_id"]))

    def create_manual_review_session(
        self,
        telegram_user_id: int,
        scheduled_for: date,
        *,
        question_limit: int | None = None,
    ) -> dict[str, Any] | None:
        plan = self.get_plan_by_telegram_id(telegram_user_id)
        if plan is None or not plan["onboarding_completed"]:
            return None

        existing_open_session = self.get_latest_open_manual_review_session(telegram_user_id)
        if existing_open_session is not None:
            return existing_open_session

        review_question_limit = int(plan["question_count"]) if question_limit is None else int(question_limit)
        review_question_limit = max(1, min(MAX_QUESTION_COUNT, review_question_limit))
        selected_questions = self._select_mistake_questions_for_user(plan["user_id"], review_question_limit)
        if not selected_questions:
            return None

        scheduled_prefix = scheduled_for.isoformat()
        review_index = 1
        while True:
            scheduled_for_text = f"{scheduled_prefix}{MANUAL_REVIEW_SESSION_MARKER}{review_index}"
            existing = self.connection.execute(
                "SELECT id FROM study_sessions WHERE user_id = ? AND scheduled_for = ?",
                (plan["user_id"], scheduled_for_text),
            ).fetchone()
            if existing is None:
                break
            review_index += 1

        return self._create_session_record(plan["user_id"], scheduled_for_text, "review", selected_questions)

    def _create_session_record(
        self,
        user_id: int,
        scheduled_for_text: str,
        session_kind: str,
        selected_questions: list[sqlite3.Row],
    ) -> dict[str, Any]:

        timestamp = utc_now_iso()
        with self._lock, self.connection:
            cursor = self.connection.execute(
                """
                INSERT INTO study_sessions (
                    user_id,
                    scheduled_for,
                    session_kind,
                    status,
                    question_count,
                    current_position,
                    created_at
                ) VALUES (?, ?, ?, 'pending', ?, 1, ?)
                """,
                (user_id, scheduled_for_text, session_kind, len(selected_questions), timestamp),
            )
            session_id = int(cursor.lastrowid)
            self.connection.executemany(
                """
                INSERT INTO session_questions (
                    session_id,
                    question_id,
                    position,
                    correct_option
                ) VALUES (?, ?, ?, ?)
                """,
                [
                    (session_id, int(item["id"]), position, str(item["correct_option"]))
                    for position, item in enumerate(selected_questions, start=1)
                ],
            )

        session_row = self.connection.execute(
            """
            SELECT ss.*, u.telegram_user_id, u.display_name
            FROM study_sessions ss
            JOIN users u ON u.id = ss.user_id
            WHERE ss.id = ?
            """,
            (session_id,),
        ).fetchone()
        return self._session_row_to_dict(session_row)

    def get_session_for_user_and_date(self, telegram_user_id: int, scheduled_for: date) -> dict[str, Any] | None:
        row = self.connection.execute(
            """
            SELECT ss.*, u.telegram_user_id, u.display_name
            FROM study_sessions ss
            JOIN users u ON u.id = ss.user_id
            WHERE u.telegram_user_id = ? AND ss.scheduled_for = ?
            """,
            (telegram_user_id, scheduled_for.isoformat()),
        ).fetchone()
        if row is None:
            return None
        return self._session_row_to_dict(row)

    def get_latest_open_manual_review_session(self, telegram_user_id: int) -> dict[str, Any] | None:
        row = self.connection.execute(
            """
            SELECT ss.*, u.telegram_user_id, u.display_name
            FROM study_sessions ss
            JOIN users u ON u.id = ss.user_id
            WHERE u.telegram_user_id = ?
              AND ss.status IN ('pending', 'in_progress')
              AND ss.session_kind = 'review'
              AND ss.scheduled_for LIKE ?
            ORDER BY ss.scheduled_for DESC
            LIMIT 1
            """,
            (telegram_user_id, f"%{MANUAL_REVIEW_SESSION_MARKER}%"),
        ).fetchone()
        if row is None:
            return None
        return self._session_row_to_dict(row)

    def get_latest_open_session_for_user(self, telegram_user_id: int) -> dict[str, Any] | None:
        row = self.connection.execute(
            """
            SELECT ss.*, u.telegram_user_id, u.display_name
            FROM study_sessions ss
            JOIN users u ON u.id = ss.user_id
            WHERE u.telegram_user_id = ? AND ss.status IN ('pending', 'in_progress')
            ORDER BY ss.scheduled_for DESC
            LIMIT 1
            """,
            (telegram_user_id,),
        ).fetchone()
        if row is None:
            return None
        return self._session_row_to_dict(row)

    def get_session_by_id(self, session_id: int) -> dict[str, Any] | None:
        row = self.connection.execute(
            """
            SELECT ss.*, u.telegram_user_id, u.display_name
            FROM study_sessions ss
            JOIN users u ON u.id = ss.user_id
            WHERE ss.id = ?
            """,
            (session_id,),
        ).fetchone()
        if row is None:
            return None
        return self._session_row_to_dict(row)

    def _get_session_question_ids(self, session_id: int) -> list[int]:
        rows = self.connection.execute(
            "SELECT question_id FROM session_questions WHERE session_id = ?",
            (session_id,),
        ).fetchall()
        return [int(row["question_id"]) for row in rows]

    def mark_session_delivered(self, session_id: int) -> None:
        with self._lock, self.connection:
            self.connection.execute(
                "UPDATE study_sessions SET delivered_at = COALESCE(delivered_at, ?) WHERE id = ?",
                (utc_now_iso(), session_id),
            )

    def mark_session_started(self, session_id: int) -> None:
        with self._lock, self.connection:
            self.connection.execute(
                """
                UPDATE study_sessions
                SET status = CASE WHEN status = 'pending' THEN 'in_progress' ELSE status END,
                    started_at = COALESCE(started_at, ?)
                WHERE id = ?
                """,
                (utc_now_iso(), session_id),
            )

    def mark_session_completed(self, session_id: int) -> None:
        with self._lock, self.connection:
            self.connection.execute(
                """
                UPDATE study_sessions
                SET status = 'completed',
                    completed_at = COALESCE(completed_at, ?)
                WHERE id = ?
                """,
                (utc_now_iso(), session_id),
            )

    def set_session_view_message_id(self, session_id: int, message_id: int | None) -> None:
        with self._lock, self.connection:
            self.connection.execute(
                "UPDATE study_sessions SET view_message_id = ? WHERE id = ?",
                (message_id, session_id),
            )

    def set_session_notice_message_id(self, session_id: int, message_id: int | None) -> None:
        with self._lock, self.connection:
            self.connection.execute(
                "UPDATE study_sessions SET notice_message_id = ? WHERE id = ?",
                (message_id, session_id),
            )

    def set_active_poll(
        self,
        session_id: int,
        *,
        poll_id: str,
        poll_message_id: int,
        session_question_id: int,
    ) -> None:
        with self._lock, self.connection:
            self.connection.execute(
                """
                UPDATE study_sessions
                SET active_poll_id = ?, active_poll_message_id = ?, active_poll_question_id = ?
                WHERE id = ?
                """,
                (poll_id, poll_message_id, session_question_id, session_id),
            )

    def clear_active_poll(self, session_id: int) -> None:
        with self._lock, self.connection:
            self.connection.execute(
                """
                UPDATE study_sessions
                SET active_poll_id = NULL, active_poll_message_id = NULL, active_poll_question_id = NULL
                WHERE id = ?
                """,
                (session_id,),
            )

    def get_active_poll_context(self, poll_id: str) -> dict[str, Any] | None:
        row = self.connection.execute(
            """
            SELECT
                ss.id AS session_id,
                ss.active_poll_question_id AS session_question_id,
                ss.active_poll_message_id,
                ss.view_message_id,
                ss.user_id,
                u.telegram_user_id
            FROM study_sessions ss
            JOIN users u ON u.id = ss.user_id
            WHERE ss.active_poll_id = ?
            """,
            (poll_id,),
        ).fetchone()
        if row is None:
            return None
        return dict(row)

    def set_session_current_position(self, session_id: int, position: int) -> None:
        max_position = int(
            self.connection.execute(
                "SELECT COUNT(*) FROM session_questions WHERE session_id = ?",
                (session_id,),
            ).fetchone()[0]
        )
        if max_position <= 0:
            return
        normalized_position = max(1, min(position, max_position))

        with self._lock, self.connection:
            self.connection.execute(
                "UPDATE study_sessions SET current_position = ? WHERE id = ?",
                (normalized_position, session_id),
            )

    def get_next_unanswered_question(self, session_id: int) -> dict[str, Any] | None:
        row = self.connection.execute(
            """
            SELECT
                sq.id AS session_question_id,
                sq.session_id,
                sq.position,
                sq.correct_option,
                sq.selected_option,
                sq.delivery_message_id,
                q.id AS question_id,
                q.caption,
                q.topic,
                q.difficulty,
                q.image_path,
                ss.question_count,
                ss.user_id,
                u.telegram_user_id
            FROM session_questions sq
            JOIN questions q ON q.id = sq.question_id
            JOIN study_sessions ss ON ss.id = sq.session_id
            JOIN users u ON u.id = ss.user_id
            WHERE sq.session_id = ? AND sq.selected_option IS NULL
            ORDER BY sq.position ASC
            LIMIT 1
            """,
            (session_id,),
        ).fetchone()
        if row is None:
            return None
        return dict(row)

    def get_session_question_by_position(self, session_id: int, position: int) -> dict[str, Any] | None:
        row = self.connection.execute(
            """
            SELECT
                sq.id AS session_question_id,
                sq.session_id,
                sq.position,
                sq.question_id,
                sq.correct_option,
                sq.selected_option,
                sq.is_correct,
                q.caption,
                q.topic,
                q.difficulty,
                q.image_path,
                ss.question_count,
                ss.current_position,
                ss.session_kind,
                ss.status,
                ss.user_id,
                u.telegram_user_id,
                u.display_name
            FROM session_questions sq
            JOIN questions q ON q.id = sq.question_id
            JOIN study_sessions ss ON ss.id = sq.session_id
            JOIN users u ON u.id = ss.user_id
            WHERE sq.session_id = ? AND sq.position = ?
            """,
            (session_id, position),
        ).fetchone()
        if row is None:
            return None
        return dict(row)

    def list_session_question_states(self, session_id: int) -> list[dict[str, Any]]:
        rows = self.connection.execute(
            """
            SELECT position, selected_option, is_correct
            FROM session_questions
            WHERE session_id = ?
            ORDER BY position ASC
            """,
            (session_id,),
        ).fetchall()
        results: list[dict[str, Any]] = []
        for row in rows:
            if row["selected_option"] is None:
                status = "unanswered"
            elif int(row["is_correct"] or 0) == 1:
                status = "correct"
            else:
                status = "wrong"
            results.append({"position": int(row["position"]), "status": status})
        return results

    def get_session_progress(self, session_id: int) -> dict[str, int]:
        row = self.connection.execute(
            """
            SELECT
                COUNT(*) AS total,
                SUM(CASE WHEN selected_option IS NULL THEN 1 ELSE 0 END) AS unanswered,
                SUM(CASE WHEN is_correct = 1 THEN 1 ELSE 0 END) AS correct,
                SUM(CASE WHEN selected_option IS NOT NULL AND COALESCE(is_correct, 0) = 0 THEN 1 ELSE 0 END) AS wrong
            FROM session_questions
            WHERE session_id = ?
            """,
            (session_id,),
        ).fetchone()
        return {
            "total": int(row["total"] or 0),
            "unanswered": int(row["unanswered"] or 0),
            "correct": int(row["correct"] or 0),
            "wrong": int(row["wrong"] or 0),
        }

    def set_delivery_message_id(self, session_question_id: int, message_id: int) -> None:
        with self._lock, self.connection:
            self.connection.execute(
                "UPDATE session_questions SET delivery_message_id = ? WHERE id = ?",
                (message_id, session_question_id),
            )

    def record_answer(self, session_question_id: int, selected_option: str) -> dict[str, Any] | None:
        row = self.connection.execute(
            """
            SELECT
                sq.id AS session_question_id,
                sq.session_id,
                sq.position,
                sq.question_id,
                sq.correct_option,
                sq.selected_option,
                ss.question_count,
                ss.current_position,
                ss.session_kind,
                u.telegram_user_id,
                ss.user_id,
                u.display_name
            FROM session_questions sq
            JOIN study_sessions ss ON ss.id = sq.session_id
            JOIN users u ON u.id = ss.user_id
            WHERE sq.id = ?
            """,
            (session_question_id,),
        ).fetchone()
        if row is None:
            return None

        context = dict(row)
        previous_selected_option = context["selected_option"]
        if previous_selected_option == selected_option:
            context["already_answered"] = True
            return context

        is_correct = selected_option == context["correct_option"]
        timestamp = utc_now_iso()
        with self._lock, self.connection:
            self.connection.execute(
                """
                UPDATE session_questions
                SET selected_option = ?, is_correct = ?, answered_at = ?
                WHERE id = ?
                """,
                (selected_option, 1 if is_correct else 0, timestamp, session_question_id),
            )
            self.connection.execute(
                """
                UPDATE study_sessions
                SET status = CASE WHEN status = 'pending' THEN 'in_progress' ELSE status END,
                    started_at = COALESCE(started_at, ?)
                WHERE id = ?
                """,
                (timestamp, context["session_id"]),
            )

        self._update_mistake_bank(
            int(context["user_id"]),
            int(context["question_id"]),
            is_correct=is_correct,
            answered_at=timestamp,
        )

        remaining = int(
            self.connection.execute(
                "SELECT COUNT(*) FROM session_questions WHERE session_id = ? AND selected_option IS NULL",
                (context["session_id"],),
            ).fetchone()[0]
        )
        correct_count = int(
            self.connection.execute(
                "SELECT COUNT(*) FROM session_questions WHERE session_id = ? AND is_correct = 1",
                (context["session_id"],),
            ).fetchone()[0]
        )

        context.update(
            {
                "previous_selected_option": previous_selected_option,
                "had_previous_answer": previous_selected_option is not None,
                "selected_option": selected_option,
                "is_correct": is_correct,
                "remaining": remaining,
                "correct_count": correct_count,
                "already_answered": False,
            }
        )
        return context

    def get_mistake_bank_summary(self, telegram_user_id: int, limit: int = 12) -> dict[str, Any] | None:
        plan = self.get_plan_by_telegram_id(telegram_user_id)
        if plan is None:
            return None

        counts_row = self.connection.execute(
            """
            SELECT
                COUNT(*) AS total,
                SUM(CASE WHEN is_active = 1 THEN 1 ELSE 0 END) AS active
            FROM mistake_bank
            WHERE user_id = ?
            """,
            (plan["user_id"],),
        ).fetchone()
        items = self.connection.execute(
            """
            SELECT q.id AS question_id, q.topic, q.difficulty, mb.wrong_count, mb.correct_count, mb.is_active
            FROM mistake_bank mb
            JOIN questions q ON q.id = mb.question_id
            WHERE mb.user_id = ? AND mb.is_active = 1 AND q.is_active = 1
            ORDER BY mb.last_wrong_at DESC
            LIMIT ?
            """,
            (plan["user_id"], limit),
        ).fetchall()
        return {
            "total": int(counts_row["total"] or 0),
            "active": int(counts_row["active"] or 0),
            "items": [
                {
                    "question_id": int(row["question_id"]),
                    "topic": str(row["topic"] or "عام"),
                    "difficulty": str(row["difficulty"] or "متوسط"),
                    "wrong_count": int(row["wrong_count"] or 0),
                    "correct_count": int(row["correct_count"] or 0),
                    "is_active": bool(row["is_active"]),
                }
                for row in items
            ],
        }

    def _update_mistake_bank(
        self,
        user_id: int,
        question_id: int,
        *,
        is_correct: bool,
        answered_at: str,
    ) -> None:
        existing = self.connection.execute(
            "SELECT user_id, question_id FROM mistake_bank WHERE user_id = ? AND question_id = ?",
            (user_id, question_id),
        ).fetchone()

        with self._lock, self.connection:
            if is_correct:
                if existing is None:
                    return
                self.connection.execute(
                    """
                    UPDATE mistake_bank
                    SET correct_count = correct_count + 1,
                        last_correct_at = ?,
                        is_active = 0
                    WHERE user_id = ? AND question_id = ?
                    """,
                    (answered_at, user_id, question_id),
                )
                return

            if existing is None:
                self.connection.execute(
                    """
                    INSERT INTO mistake_bank (
                        user_id,
                        question_id,
                        first_wrong_at,
                        last_wrong_at,
                        wrong_count,
                        correct_count,
                        last_correct_at,
                        is_active
                    ) VALUES (?, ?, ?, ?, 1, 0, NULL, 1)
                    """,
                    (user_id, question_id, answered_at, answered_at),
                )
                return

            self.connection.execute(
                """
                UPDATE mistake_bank
                SET wrong_count = wrong_count + 1,
                    last_wrong_at = ?,
                    is_active = 1
                WHERE user_id = ? AND question_id = ?
                """,
                (answered_at, user_id, question_id),
            )

    def _select_mistake_questions_for_user(
        self,
        user_id: int,
        count: int,
        *,
        exclude_question_ids: list[int] | None = None,
    ) -> list[sqlite3.Row]:
        if count <= 0:
            return []

        excluded_ids = [int(question_id) for question_id in (exclude_question_ids or [])]
        exclusion_clause = ""
        parameters: list[Any] = [user_id]
        if excluded_ids:
            placeholders = ",".join("?" for _ in excluded_ids)
            exclusion_clause = f" AND q.id NOT IN ({placeholders})"
            parameters.extend(excluded_ids)
        parameters.append(count)

        rows = self.connection.execute(
            f"""
            SELECT q.id, q.correct_option
            FROM mistake_bank mb
            JOIN questions q ON q.id = mb.question_id
            WHERE mb.user_id = ? AND mb.is_active = 1 AND q.is_active = 1 AND q.correct_option IN ('A', 'B', 'C', 'D')
            {exclusion_clause}
            ORDER BY COALESCE(mb.last_correct_at, '') ASC, mb.last_wrong_at DESC
            LIMIT ?
            """,
            tuple(parameters),
        ).fetchall()
        return list(rows)

    @staticmethod
    def _determine_session_kind(plan: dict[str, Any], scheduled_for: date) -> str:
        review_weekday = plan.get("review_weekday")
        if review_weekday is None:
            return "study"
        if len(plan["weekdays"]) <= 1:
            return "study"
        if scheduled_for.weekday() == int(review_weekday):
            return "review"
        return "study"

    def get_session_result(self, session_id: int) -> dict[str, int]:
        total = int(
            self.connection.execute(
                "SELECT COUNT(*) FROM session_questions WHERE session_id = ?",
                (session_id,),
            ).fetchone()[0]
        )
        correct = int(
            self.connection.execute(
                "SELECT COUNT(*) FROM session_questions WHERE session_id = ? AND is_correct = 1",
                (session_id,),
            ).fetchone()[0]
        )
        return {"total": total, "correct": correct}

    def _expire_old_sessions(self, user_id: int, scheduled_for_text: str) -> None:
        with self._lock, self.connection:
            self.connection.execute(
                """
                UPDATE study_sessions
                SET status = 'expired'
                WHERE user_id = ? AND scheduled_for < ? AND status IN ('pending', 'in_progress')
                """,
                (user_id, scheduled_for_text),
            )

    def _select_questions_for_plan(
        self,
        plan: dict[str, Any],
        scheduled_for: date,
    ) -> tuple[str, list[sqlite3.Row]]:
        session_kind = self._determine_session_kind(plan, scheduled_for)
        if session_kind == "review":
            selected_questions = self._select_mistake_questions_for_user(
                int(plan["user_id"]),
                int(plan["question_count"]),
            )
            if not selected_questions:
                session_kind = "study"
                selected_questions = self._select_questions_for_user(
                    int(plan["user_id"]),
                    int(plan["question_count"]),
                )
        else:
            selected_questions = self._select_questions_for_user(
                int(plan["user_id"]),
                int(plan["question_count"]),
            )
        return session_kind, selected_questions

    def _select_questions_for_user(
        self,
        user_id: int,
        count: int,
        *,
        exclude_question_ids: list[int] | None = None,
    ) -> list[sqlite3.Row]:
        if count <= 0:
            return []

        excluded_ids = [int(question_id) for question_id in (exclude_question_ids or [])]
        fresh_exclusion_clause = ""
        fresh_parameters: list[Any] = []
        if excluded_ids:
            placeholders = ",".join("?" for _ in excluded_ids)
            fresh_exclusion_clause = f" AND q.id NOT IN ({placeholders})"
            fresh_parameters.extend(excluded_ids)
        fresh_parameters.extend([user_id, count])

        fresh_rows = self.connection.execute(
            f"""
            SELECT q.id, q.correct_option
            FROM questions q
                        WHERE q.is_active = 1
                            AND q.correct_option IN ('A', 'B', 'C', 'D')
              {fresh_exclusion_clause}
              AND NOT EXISTS (
                  SELECT 1
                  FROM session_questions sq
                  JOIN study_sessions ss ON ss.id = sq.session_id
                  WHERE ss.user_id = ? AND sq.question_id = q.id
              )
            ORDER BY RANDOM()
            LIMIT ?
            """,
            tuple(fresh_parameters),
        ).fetchall()
        if len(fresh_rows) >= count:
            return list(fresh_rows)

        selected_ids = excluded_ids + [int(row["id"]) for row in fresh_rows]
        exclusion_clause = ""
        parameters: list[Any] = []
        if selected_ids:
            placeholders = ",".join("?" for _ in selected_ids)
            exclusion_clause = f" AND q.id NOT IN ({placeholders})"
            parameters.extend(selected_ids)
        parameters.extend([user_id, count - len(fresh_rows)])

        fallback_rows = self.connection.execute(
            f"""
            SELECT q.id, q.correct_option
            FROM questions q
            WHERE q.is_active = 1 AND q.correct_option IN ('A', 'B', 'C', 'D') {exclusion_clause}
            ORDER BY COALESCE((
                SELECT MAX(ss.scheduled_for)
                FROM session_questions sq
                JOIN study_sessions ss ON ss.id = sq.session_id
                WHERE ss.user_id = ? AND sq.question_id = q.id
            ), '') ASC, RANDOM()
            LIMIT ?
            """,
            tuple(parameters),
        ).fetchall()

        return list(fresh_rows) + list(fallback_rows)

    def _question_row_to_public(self, row: sqlite3.Row) -> dict[str, Any]:
        normalized_option = str(row["correct_option"] or "").strip().upper()
        return {
            "id": int(row["id"]),
            "question_number": int(row["question_number"] or row["id"]),
            "caption": str(row["caption"] or ""),
            "topic": str(row["topic"] or "عام"),
            "difficulty": str(row["difficulty"] or "متوسط"),
            "correct_option": normalized_option,
            "is_ready": normalized_option in VALID_OPTIONS,
            "created_at": str(row["created_at"]),
            "image_url": f"/uploads/{Path(str(row['image_path'])).name}",
            "assignment_count": int(row["assignment_count"] or 0),
        }

    @staticmethod
    def _plan_row_to_dict(row: sqlite3.Row) -> dict[str, Any]:
        weekdays_raw = str(row["weekdays"] or "")
        weekdays = sorted(int(item) for item in weekdays_raw.split(",") if item != "")
        review_weekday = row["review_weekday"]
        return {
            "plan_id": int(row["plan_id"]),
            "user_id": int(row["user_id"]),
            "telegram_user_id": int(row["telegram_user_id"]),
            "display_name": str(row["display_name"]),
            "reminder_time": str(row["reminder_time"]),
            "question_count": int(row["question_count"]),
            "review_weekday": int(review_weekday) if review_weekday is not None else None,
            "timezone_name": str(row["timezone_name"]),
            "is_active": bool(row["is_active"]),
            "onboarding_completed": bool(row["onboarding_completed"]),
            "weekdays": weekdays,
        }

    @staticmethod
    def _session_row_to_dict(row: sqlite3.Row) -> dict[str, Any]:
        return {
            "session_id": int(row["id"]),
            "user_id": int(row["user_id"]),
            "telegram_user_id": int(row["telegram_user_id"]),
            "display_name": str(row["display_name"]),
            "scheduled_for": str(row["scheduled_for"]),
            "session_kind": str(row["session_kind"] or "study"),
            "status": str(row["status"]),
            "question_count": int(row["question_count"]),
            "current_position": int(row["current_position"] or 1),
            "notice_message_id": row["notice_message_id"],
            "view_message_id": row["view_message_id"],
            "active_poll_id": row["active_poll_id"],
            "active_poll_message_id": row["active_poll_message_id"],
            "active_poll_question_id": row["active_poll_question_id"],
            "delivered_at": row["delivered_at"],
            "started_at": row["started_at"],
            "completed_at": row["completed_at"],
        }

    @staticmethod
    def _resolve_suffix(original_filename: str) -> str:
        suffix = Path(original_filename or "").suffix.lower()
        if suffix in VALID_IMAGE_SUFFIXES:
            return suffix
        return ".jpg"


def utc_now_iso() -> str:
    return datetime.utcnow().replace(microsecond=0).isoformat() + "Z"


def _collapse_repeated_env_value(value: str) -> str:
    normalized = value.strip()
    if not normalized:
        return normalized

    length = len(normalized)
    for part_length in range(1, (length // 2) + 1):
        if length % part_length != 0:
            continue
        part = normalized[:part_length]
        if part * (length // part_length) == normalized:
            return part
    return normalized


def _normalize_host_env(value: str, *, default: str) -> str:
    normalized = _collapse_repeated_env_value(value).strip()
    if not normalized:
        return default

    if re.fullmatch(r"\d{1,3}(?:\.\d{1,3}){3}", normalized):
        return normalized

    overlap_matches = re.finditer(r"(?=(\d{1,3}(?:\.\d{1,3}){3}))", normalized)
    valid_matches = []
    for match in overlap_matches:
        candidate = match.group(1)
        octets = candidate.split(".")
        if all(0 <= int(octet) <= 255 for octet in octets):
            valid_matches.append(candidate)
    if valid_matches:
        return valid_matches[-1]

    return normalized


def _normalize_port_env(value: str, *, fallback: str) -> str:
    normalized = _collapse_repeated_env_value(value).strip()
    if not normalized:
        return fallback

    if normalized.isdigit():
        try:
            port_value = int(normalized)
        except ValueError:
            port_value = -1
        if 0 < port_value <= 65535:
            return normalized

    candidate_values = [fallback.strip()]
    if fallback.strip().isdigit() and fallback.strip() in normalized:
        candidate_values.insert(0, fallback.strip())

    for candidate in candidate_values:
        if candidate.isdigit() and 0 < int(candidate) <= 65535:
            return candidate

    for suffix_length in range(5, 0, -1):
        if len(normalized) < suffix_length:
            continue
        candidate = normalized[-suffix_length:]
        if not candidate.isdigit():
            continue
        port_value = int(candidate)
        if 0 < port_value <= 65535:
            return candidate

    raise RuntimeError("ABQOOR_PORT must be a valid integer.")


def _parse_boolean_env(value: str, *, default: bool) -> bool:
    normalized = _collapse_repeated_env_value(value).strip().lower()
    if not normalized:
        return default
    if normalized in {"1", "true", "yes", "on"}:
        return True
    if normalized in {"0", "false", "no", "off"}:
        return False
    if set(normalized) <= {"0", "1"}:
        return normalized[-1] == "1"
    return default


def load_settings() -> Settings:
    load_dotenv()

    telegram_bot_token = os.getenv("TELEGRAM_BOT_TOKEN", "").strip()
    admin_password = os.getenv("ADMIN_PASSWORD", "").strip()
    host_default = os.getenv("HOST", "127.0.0.1")
    port_default = os.getenv("PORT", "8000")
    host = _normalize_host_env(os.getenv("ABQOOR_HOST", host_default), default=host_default.strip() or "127.0.0.1")
    port_value = _normalize_port_env(os.getenv("ABQOOR_PORT", port_default), fallback=port_default.strip() or "8000")
    storage_dir_value = _collapse_repeated_env_value(os.getenv("ABQOOR_STORAGE_DIR", str(BASE_DIR))) or str(BASE_DIR)
    timezone_name = _collapse_repeated_env_value(os.getenv("ABQOOR_TIMEZONE", "Asia/Riyadh")) or "Asia/Riyadh"
    telegram_enabled_value = os.getenv("ABQOOR_ENABLE_TELEGRAM", "1").strip().lower()
    telegram_enabled = _parse_boolean_env(telegram_enabled_value, default=True)

    missing = []
    if telegram_enabled and not telegram_bot_token:
        missing.append("TELEGRAM_BOT_TOKEN")
    if not admin_password:
        missing.append("ADMIN_PASSWORD")
    if missing:
        raise RuntimeError("Missing required environment variables: " + ", ".join(missing))

    port = int(port_value)

    try:
        ZoneInfo(timezone_name)
    except ZoneInfoNotFoundError as exc:
        raise RuntimeError("ABQOOR_TIMEZONE must be a valid IANA timezone name.") from exc

    return Settings(
        telegram_bot_token=telegram_bot_token,
        admin_password=admin_password,
        host=host,
        port=port,
        storage_dir=Path(storage_dir_value).expanduser().resolve(),
        timezone_name=timezone_name,
        telegram_enabled=telegram_enabled,
    )


def validate_admin_token(settings: Settings, admin_token: str | None) -> None:
    if admin_token != settings.admin_password:
        raise HTTPException(status_code=401, detail="كلمة مرور المشرف غير صحيحة.")


def build_bot_deep_link(bot_username: str | None, payload: str = CHANNEL_START_PAYLOAD) -> str | None:
    if not bot_username:
        return None
    return create_deep_linked_url(bot_username, payload)


def normalize_user_input(value: str) -> str:
    return value.translate(ARABIC_DIGITS_TRANSLATION).strip()


def parse_positive_integer_input(raw_value: str) -> int | None:
    normalized = normalize_user_input(raw_value)
    if not normalized.isdigit():
        return None

    value = int(normalized)
    if value <= 0:
        return None
    return value


def build_question_number_conflict_detail(conflicts: list[dict[str, Any]]) -> dict[str, Any]:
    sorted_conflicts = sorted(conflicts, key=lambda item: int(item["question_number"]))
    question_numbers = [int(item["question_number"]) for item in sorted_conflicts]
    clipboard_text = "\n".join(str(number) for number in question_numbers)

    if len(sorted_conflicts) == 1:
        question_number = question_numbers[0]
        return {
            "message": f"رقم السؤال {question_number} موجود بالفعل. اختر هل تريد تخطيه أم استبدال السؤال الحالي.",
            "mode": "single",
            "question_numbers": question_numbers,
            "clipboard_text": clipboard_text,
            "conflict_count": 1,
            "conflict_question": sorted_conflicts[0],
        }

    return {
        "message": (
            f"وجدت {len(question_numbers)} أرقام أسئلة موجودة بالفعل. "
            "نسخت الأرقام إلى الحافظة؛ اختر التخطي أو الاستبدال للجميع."
        ),
        "mode": "multiple",
        "question_numbers": question_numbers,
        "clipboard_text": clipboard_text,
        "conflict_count": len(question_numbers),
    }


def normalize_spreadsheet_token(value: Any) -> str:
    normalized = normalize_user_input(str(value or "")).lower().strip()
    normalized = normalized.replace("أ", "ا").replace("إ", "ا").replace("آ", "ا")
    return "".join(character for character in normalized if character.isalnum())


def detect_spreadsheet_column(headers: tuple[Any, ...], aliases: set[str]) -> int | None:
    normalized_aliases = {normalize_spreadsheet_token(alias) for alias in aliases}
    for index, header in enumerate(headers):
        normalized_header = normalize_spreadsheet_token(header)
        if normalized_header and normalized_header in normalized_aliases:
            return index
    return None


def parse_question_number_cell(value: Any) -> int | None:
    if value is None or value == "":
        return None

    if isinstance(value, bool):
        return None
    if isinstance(value, int):
        return value if value > 0 else None
    if isinstance(value, float):
        if value.is_integer() and value > 0:
            return int(value)
        return None

    normalized = normalize_user_input(str(value))
    if not normalized:
        return None
    if normalized.isdigit():
        parsed = int(normalized)
        return parsed if parsed > 0 else None

    try:
        numeric_value = float(normalized)
    except ValueError:
        numeric_value = None
    if numeric_value is not None and numeric_value.is_integer() and numeric_value > 0:
        return int(numeric_value)

    match = re.search(r"\d+", normalized)
    if match is None:
        return None

    parsed = int(match.group(0))
    return parsed if parsed > 0 else None


def parse_answer_option_cell(value: Any) -> str:
    if value is None or value == "":
        return ""

    if isinstance(value, bool):
        return ""
    if isinstance(value, int):
        if 1 <= value <= 4:
            return POLL_OPTION_ORDER[value - 1]
        return ""
    if isinstance(value, float):
        if value.is_integer() and 1 <= int(value) <= 4:
            return POLL_OPTION_ORDER[int(value) - 1]
        return ""

    normalized = normalize_spreadsheet_token(value)
    option_aliases = {
        "A": {"a", "1", "ا"},
        "B": {"b", "2", "ب"},
        "C": {"c", "3", "ج"},
        "D": {"d", "4", "د"},
    }
    for option, aliases in option_aliases.items():
        normalized_aliases = {normalize_spreadsheet_token(alias) for alias in aliases}
        if normalized in normalized_aliases:
            return option
    return ""


def extract_answer_key_from_sheet(filename: str, sheet_path: Path) -> dict[int, str]:
    suffix = Path(filename or "").suffix.lower()
    if suffix not in VALID_ANSWER_SHEET_SUFFIXES:
        raise HTTPException(
            status_code=400,
            detail="ملف الإجابات يجب أن يكون Excel بصيغة XLSX أو XLSM، أو CSV.",
        )

    def extract_answer_key_from_rows(headers: tuple[Any, ...], rows: Any) -> dict[int, str]:
        question_number_column = detect_spreadsheet_column(headers, QUESTION_NUMBER_HEADER_ALIASES)
        answer_column = detect_spreadsheet_column(headers, ANSWER_HEADER_ALIASES)

        if question_number_column is None or answer_column is None:
            raise HTTPException(
                status_code=400,
                detail=(
                    "لم أتمكن من العثور على عمود رقم السؤال وعمود الإجابة في ملف Excel. "
                    "أرسل شكل الملف إذا كانت أسماء الأعمدة مختلفة وسأضبط القراءة عليه."
                ),
            )

        answer_key: dict[int, str] = {}
        for row in rows:
            question_number = parse_question_number_cell(row[question_number_column] if question_number_column < len(row) else None)
            correct_option = parse_answer_option_cell(row[answer_column] if answer_column < len(row) else None)
            if question_number is None or not correct_option:
                continue
            answer_key[question_number] = correct_option

        if not answer_key:
            raise HTTPException(
                status_code=400,
                detail="لم أجد صفوفًا صالحة تحتوي على رقم سؤال وإجابة يمكن استخدامها.",
            )

        return answer_key

    if suffix == ".csv":
        last_decode_error: UnicodeDecodeError | None = None
        for encoding in ("utf-8-sig", "utf-8", "cp1256"):
            try:
                with sheet_path.open("r", encoding=encoding, newline="") as csv_file:
                    reader = csv.reader(csv_file)
                    headers_row = next(reader, None)
                    if headers_row is None:
                        raise HTTPException(status_code=400, detail="ملف الإجابات فارغ.")
                    return extract_answer_key_from_rows(tuple(headers_row), reader)
            except UnicodeDecodeError as exc:
                last_decode_error = exc
                continue
        if last_decode_error is not None:
            raise HTTPException(status_code=400, detail="تعذر قراءة ملف CSV المرفوع.")
        raise HTTPException(status_code=400, detail="ملف الإجابات فارغ.")

    workbook = None
    try:
        workbook = load_workbook(filename=sheet_path, read_only=True, data_only=True)
        worksheet = workbook.active
        rows_iterator = worksheet.iter_rows(values_only=True)
        headers = next(rows_iterator, None)
        if headers is None:
            raise HTTPException(status_code=400, detail="ملف Excel فارغ.")
        return extract_answer_key_from_rows(tuple(headers), rows_iterator)
    except HTTPException:
        raise
    except Exception as exc:
        raise HTTPException(status_code=400, detail="تعذر قراءة ملف Excel المرفوع.") from exc
    finally:
        if workbook is not None:
            workbook.close()


def read_pdf_page_count(pdf_path: Path) -> int:
    try:
        document = fitz.open(str(pdf_path))
    except Exception as exc:
        raise HTTPException(status_code=400, detail="تعذر قراءة ملف PDF المرفوع.") from exc

    try:
        if document.needs_pass and not document.authenticate(""):
            raise HTTPException(status_code=400, detail="ملف PDF محمي بكلمة مرور ولا يمكن تحويله إلى صور.")
        return int(document.page_count)
    finally:
        document.close()


async def save_upload_to_temporary_path(upload: UploadFile, suffix: str) -> tuple[Path, int]:
    temporary_file = tempfile.NamedTemporaryFile(delete=False, suffix=suffix)
    temporary_path = Path(temporary_file.name)
    bytes_written = 0

    try:
        with temporary_file:
            await upload.seek(0)
            while True:
                chunk = await upload.read(PDF_UPLOAD_CHUNK_SIZE)
                if not chunk:
                    break
                temporary_file.write(chunk)
                bytes_written += len(chunk)
        await upload.seek(0)
        return temporary_path, bytes_written
    except Exception:
        temporary_path.unlink(missing_ok=True)
        raise


def parse_time_input(raw_value: str) -> str | None:
    normalized = normalize_user_input(raw_value).replace(" ", "")
    if not normalized:
        return None

    if ":" not in normalized:
        if not normalized.isdigit():
            return None
        hour = int(normalized)
        minute = 0
    else:
        parts = normalized.split(":")
        if len(parts) != 2 or not parts[0].isdigit() or not parts[1].isdigit():
            return None
        hour = int(parts[0])
        minute = int(parts[1])

    if hour < 0 or hour > 23 or minute < 0 or minute > 59:
        return None
    return f"{hour:02d}:{minute:02d}"


def parse_question_count_input(raw_value: str) -> int | None:
    normalized = normalize_user_input(raw_value)
    if not normalized.isdigit():
        return None

    value = int(normalized)
    if value < MIN_QUESTION_COUNT or value > MAX_QUESTION_COUNT:
        return None
    return value


def format_time_label(reminder_time: str) -> str:
    parsed = parse_time_input(reminder_time)
    if parsed is None:
        return reminder_time

    hour, minute = (int(part) for part in parsed.split(":"))
    suffix = "صباحًا" if hour < 12 else "مساءً"
    display_hour = hour % 12 or 12
    return f"{display_hour}:{minute:02d} {suffix}"


def format_weekdays(weekdays: list[int]) -> str:
    ordered = [label for weekday, label in ARABIC_WEEKDAY_ORDER if weekday in weekdays]
    return "، ".join(ordered) if ordered else "غير محددة"


def build_plan_summary(plan: dict[str, Any]) -> str:
    status_text = "مفعلة" if plan["is_active"] else "متوقفة"
    review_weekday = plan.get("review_weekday")
    if review_weekday is None:
        review_text = "غير مخصص"
    else:
        review_text = ARABIC_WEEKDAY_LABELS.get(int(review_weekday), "غير مخصص")
    return (
        f"أيام الدراسة: {format_weekdays(plan['weekdays'])}\n"
        f"وقت التذكير: {format_time_label(plan['reminder_time'])}\n"
        f"عدد الأسئلة: {plan['question_count']}\n"
        f"يوم مراجعة الأخطاء: {review_text}\n"
        f"الحالة: {status_text}"
    )
def build_menu_keyboard(plan: dict[str, Any], *, today_label: str = TODAY_START_LABEL) -> InlineKeyboardMarkup:
    toggle_label = "أوقِف التذكير" if plan["is_active"] else "فعّل التذكير"
    toggle_action = "pause" if plan["is_active"] else "resume"
    return InlineKeyboardMarkup(
        [
            [
                InlineKeyboardButton(today_label, callback_data="menu:today"),
                InlineKeyboardButton("بنك الأخطاء", callback_data="menu:mistakes"),
            ],
            [
                InlineKeyboardButton("عدّل خطتي", callback_data="menu:plan"),
                InlineKeyboardButton(toggle_label, callback_data=f"menu:{toggle_action}"),
            ],
        ]
    )


def build_session_start_keyboard() -> InlineKeyboardMarkup:
    return InlineKeyboardMarkup(
        [[InlineKeyboardButton("ابدأ", callback_data="menu:today")]]
    )


def get_today_menu_label(application: Application, telegram_user_id: int) -> str:
    repository = get_repository(application)
    settings = get_settings(application)
    today_session = repository.get_session_for_user_and_date(
        telegram_user_id,
        datetime.now(settings.timezone).date(),
    )
    if today_session is not None and today_session["status"] == "completed":
        return TODAY_REDO_LABEL
    return TODAY_START_LABEL


def build_user_menu_keyboard(application: Application, plan: dict[str, Any]) -> InlineKeyboardMarkup:
    return build_menu_keyboard(
        plan,
        today_label=get_today_menu_label(application, int(plan["telegram_user_id"])),
    )


def build_plan_setup_summary(state: dict[str, Any]) -> str:
    weekdays = sorted(state.get("days", set()))
    reminder_time = state.get("reminder_time")
    question_count = state.get("question_count")
    review_weekday = state.get("review_weekday")

    if len(weekdays) <= 1:
        review_text = "غير متاح إلا عند اختيار يومي دراسة أو أكثر"
    elif review_weekday is None:
        review_text = "بدون يوم مراجعة"
    else:
        review_text = ARABIC_WEEKDAY_LABELS.get(int(review_weekday), "بدون يوم مراجعة")

    time_text = format_time_label(str(reminder_time)) if reminder_time else "غير محدد"
    count_text = str(question_count) if question_count is not None else "غير محدد"

    return (
        "عدّل خطتك من هنا، ثم اضغط حفظ الخطة بعد الانتهاء.\n\n"
        f"أيام الدراسة: {format_weekdays(weekdays)}\n"
        f"وقت التذكير: {time_text}\n"
        f"عدد الأسئلة يوميًا: {count_text}\n"
        f"يوم مراجعة الأخطاء: {review_text}"
    )


def build_plan_edit_keyboard(state: dict[str, Any]) -> InlineKeyboardMarkup:
    review_label = "يوم مراجعة الأخطاء"
    if len(state.get("days", set())) <= 1:
        review_label = "يوم المراجعة غير مفعل"
    return InlineKeyboardMarkup(
        [
            [
                InlineKeyboardButton("أيام الدراسة", callback_data="setup:edit:days"),
                InlineKeyboardButton("وقت التذكير", callback_data="setup:edit:time"),
            ],
            [
                InlineKeyboardButton("عدد الأسئلة يوميًا", callback_data="setup:edit:count"),
                InlineKeyboardButton(review_label, callback_data="setup:edit:review"),
            ],
            [
                InlineKeyboardButton("حفظ الخطة", callback_data="setup:save"),
                InlineKeyboardButton("إلغاء", callback_data="setup:cancel"),
            ],
        ]
    )


def build_days_keyboard(selected_days: set[int]) -> InlineKeyboardMarkup:
    rows: list[list[InlineKeyboardButton]] = []
    current_row: list[InlineKeyboardButton] = []
    for weekday, label in ARABIC_WEEKDAY_ORDER:
        prefix = "✓ " if weekday in selected_days else ""
        current_row.append(
            InlineKeyboardButton(f"{prefix}{label}", callback_data=f"setup:day:{weekday}")
        )
        if len(current_row) == 2:
            rows.append(current_row)
            current_row = []
    if current_row:
        rows.append(current_row)

    rows.append([InlineKeyboardButton("تم", callback_data="setup:days:done")])
    rows.append([InlineKeyboardButton("إلغاء", callback_data="setup:cancel")])
    return InlineKeyboardMarkup(rows)


def build_review_day_keyboard(selected_days: list[int], current_review_day: int | None) -> InlineKeyboardMarkup:
    rows: list[list[InlineKeyboardButton]] = []
    current_row: list[InlineKeyboardButton] = []
    for weekday in selected_days:
        label = ARABIC_WEEKDAY_LABELS.get(weekday, str(weekday))
        prefix = "✓ " if current_review_day == weekday else ""
        current_row.append(InlineKeyboardButton(f"{prefix}{label}", callback_data=f"setup:review:{weekday}"))
        if len(current_row) == 2:
            rows.append(current_row)
            current_row = []
    if current_row:
        rows.append(current_row)
    rows.append([InlineKeyboardButton("بدون يوم مراجعة", callback_data="setup:review:none")])
    rows.append([InlineKeyboardButton("إلغاء", callback_data="setup:cancel")])
    return InlineKeyboardMarkup(rows)


def build_time_hour_keyboard(selected_hour: int | None = None) -> InlineKeyboardMarkup:
    rows: list[list[InlineKeyboardButton]] = []
    current_row: list[InlineKeyboardButton] = []
    for hour in range(24):
        label = format_time_label(f"{hour:02d}:00")
        if selected_hour == hour:
            label = f"✓ {label}"
        current_row.append(InlineKeyboardButton(label, callback_data=f"setup:timehour:{hour:02d}"))
        if len(current_row) == 3:
            rows.append(current_row)
            current_row = []
    if current_row:
        rows.append(current_row)
    rows.append([InlineKeyboardButton("إلغاء", callback_data="setup:cancel")])
    return InlineKeyboardMarkup(rows)


def build_time_minute_keyboard(selected_hour: int, selected_minute: int | None = None) -> InlineKeyboardMarkup:
    rows: list[list[InlineKeyboardButton]] = []
    current_row: list[InlineKeyboardButton] = []
    for minute in (0, 15, 30, 45):
        label = format_time_label(f"{selected_hour:02d}:{minute:02d}")
        if selected_minute == minute:
            label = f"✓ {label}"
        current_row.append(InlineKeyboardButton(label, callback_data=f"setup:timeminute:{minute:02d}"))
        if len(current_row) == 2:
            rows.append(current_row)
            current_row = []
    if current_row:
        rows.append(current_row)
    rows.append(
        [
            InlineKeyboardButton("تغيير الساعة", callback_data="setup:timeback"),
            InlineKeyboardButton("إلغاء", callback_data="setup:cancel"),
        ]
    )
    return InlineKeyboardMarkup(rows)


def build_question_count_keyboard(current_count: int | None = None) -> InlineKeyboardMarkup:
    values = [1, 2, 3, 4, 5, 6, 8, 10, 12, 15]
    rows: list[list[InlineKeyboardButton]] = []
    current_row: list[InlineKeyboardButton] = []
    for value in values:
        label = f"{value}"
        if current_count == value:
            label = f"✓ {value}"
        current_row.append(InlineKeyboardButton(label, callback_data=f"setup:count:{value}"))
        if len(current_row) == 3:
            rows.append(current_row)
            current_row = []
    if current_row:
        rows.append(current_row)
    rows.append([InlineKeyboardButton("إلغاء", callback_data="setup:cancel")])
    return InlineKeyboardMarkup(rows)


def get_mistake_review_max_count(summary: dict[str, Any]) -> int:
    return min(MAX_QUESTION_COUNT, max(0, int(summary["active"])))


def build_mistake_review_bucket_values(max_review_count: int) -> list[int]:
    if max_review_count <= 0:
        return []
    bucket_values = [1]
    if max_review_count >= 10:
        bucket_values.extend(range(10, max_review_count + 1, 10))
    return bucket_values


def build_mistake_review_exact_values(bucket_start: int, max_review_count: int) -> list[int]:
    if max_review_count <= 0:
        return []
    if max_review_count < 10:
        return list(range(1, max_review_count + 1))
    if bucket_start <= 1:
        return list(range(1, min(9, max_review_count) + 1))

    start_value = max(10, bucket_start)
    if start_value >= MAX_QUESTION_COUNT:
        end_value = min(MAX_QUESTION_COUNT, max_review_count)
    else:
        end_value = min(start_value + 9, max_review_count)
    if start_value > end_value:
        return []
    return list(range(start_value, end_value + 1))


def build_mistake_review_bucket_text(*, max_review_count: int, page: int) -> str:
    lines = [
        "اختر الفئة التي يقع فيها العدد الذي تريد مراجعته الآن.",
        f"يمكنك بدء جلسة مراجعة بأي عدد حتى {max_review_count} سؤال.",
    ]
    if max_review_count > 50:
        if page <= 0:
            lines.append("إذا لم تجد الفئة المناسبة هنا، اضغط متابعة لعرض أعداد أكبر حتى 100.")
        else:
            lines.append("استخدم أزرار التنقل بالأسفل للرجوع أو متابعة تصفح بقية الفئات.")
    return "\n".join(lines)


def build_mistake_review_bucket_keyboard(*, max_review_count: int, page: int) -> InlineKeyboardMarkup:
    bucket_values = build_mistake_review_bucket_values(max_review_count)
    page_size = 6
    total_pages = max(1, (len(bucket_values) + page_size - 1) // page_size)
    safe_page = min(max(page, 0), total_pages - 1)
    start_index = safe_page * page_size
    visible_values = bucket_values[start_index:start_index + page_size]

    rows: list[list[InlineKeyboardButton]] = []
    current_row: list[InlineKeyboardButton] = []
    for value in visible_values:
        current_row.append(
            InlineKeyboardButton(
                "1-9" if value == 1 else str(value),
                callback_data=f"menu:mistakes_review_exact:{value}:{safe_page}",
            )
        )
        if len(current_row) == 3:
            rows.append(current_row)
            current_row = []
    if current_row:
        rows.append(current_row)

    nav_row: list[InlineKeyboardButton] = []
    if safe_page > 0:
        nav_row.append(InlineKeyboardButton("السابق", callback_data=f"menu:mistakes_review_pick:{safe_page - 1}"))
    if safe_page < total_pages - 1:
        nav_row.append(InlineKeyboardButton("متابعة", callback_data=f"menu:mistakes_review_pick:{safe_page + 1}"))
    if nav_row:
        rows.append(nav_row)

    rows.append([InlineKeyboardButton("رجوع إلى بنك الأخطاء", callback_data="menu:mistakes")])
    return InlineKeyboardMarkup(rows)


def build_mistake_review_exact_text(*, bucket_start: int, max_review_count: int) -> str:
    exact_values = build_mistake_review_exact_values(bucket_start, max_review_count)
    if not exact_values:
        return "لا توجد أعداد متاحة لهذه الفئة الآن."
    first_value = exact_values[0]
    last_value = exact_values[-1]
    if first_value == last_value:
        return f"اختر بدء مراجعة {first_value} سؤال الآن."
    return (
        "اختر العدد الدقيق الذي تريد مراجعته الآن.\n"
        f"هذه الفئة تغطي من {first_value} إلى {last_value} سؤال."
    )


def build_mistake_review_exact_keyboard(
    *,
    bucket_start: int,
    bucket_page: int,
    max_review_count: int,
    include_back_actions: bool = True,
) -> InlineKeyboardMarkup:
    exact_values = build_mistake_review_exact_values(bucket_start, max_review_count)
    rows: list[list[InlineKeyboardButton]] = []
    current_row: list[InlineKeyboardButton] = []
    for value in exact_values:
        current_row.append(
            InlineKeyboardButton(
                str(value),
                callback_data=f"menu:mistakes_review_start:{value}",
            )
        )
        if len(current_row) == 3:
            rows.append(current_row)
            current_row = []
    if current_row:
        rows.append(current_row)

    if include_back_actions:
        if max_review_count < 10:
            rows.append([InlineKeyboardButton("رجوع إلى بنك الأخطاء", callback_data="menu:mistakes")])
        else:
            rows.append(
                [
                    InlineKeyboardButton("رجوع إلى الفئات", callback_data=f"menu:mistakes_review_pick:{bucket_page}"),
                    InlineKeyboardButton("بنك الأخطاء", callback_data="menu:mistakes"),
                ]
            )
    return InlineKeyboardMarkup(rows)


def build_mistake_bank_text(
    summary: dict[str, Any],
    *,
    max_review_count: int | None = None,
    has_open_session: bool = False,
) -> str:
    if summary["total"] == 0:
        return "بنك الأخطاء فارغ الآن. كلما أخطأت في سؤال سأحفظه هنا حتى تراجعه لاحقًا."

    active_count = int(summary["active"])
    lines: list[str] = []
    if active_count == 0:
        lines.append("لا توجد أسئلة تحتاج مراجعة الآن. كل سؤال أجبته بشكل صحيح خرج من بنك الأخطاء.")
        lines.append(f"إجمالي ما مر على البنك حتى الآن: {summary['total']} سؤال، ولا يوجد منها شيء نشط الآن.")
    else:
        lines.append(
            f"بنك الأخطاء يحتوي الآن على {active_count} سؤال يحتاج مراجعة من أصل {summary['total']} سؤال مر على البنك."
        )
        if has_open_session:
            lines.append("لديك جلسة مراجعة مفتوحة بالفعل. يمكنك إكمالها من الزر أسفل هذه الرسالة.")
        elif max_review_count is not None and max_review_count > 0:
            if max_review_count < 10:
                lines.append("اختر العدد مباشرة من الأزرار أسفل هذه الرسالة.")
            else:
                lines.append(
                    "اختر من الأزرار أسفل هذه الرسالة فئة العدد أولًا مثل 10 أو 20 أو 30، ثم سأعرض لك الأعداد التفصيلية داخلها."
                )

    return "\n".join(lines)


def build_mistake_bank_keyboard(*, max_review_count: int, has_open_session: bool) -> InlineKeyboardMarkup | None:
    if has_open_session:
        return InlineKeyboardMarkup(
            [[InlineKeyboardButton("أكمل مراجعة الأخطاء", callback_data="menu:mistakes_review")]]
        )
    if max_review_count <= 0:
        return None
    if max_review_count < 10:
        return build_mistake_review_exact_keyboard(
            bucket_start=1,
            bucket_page=0,
            max_review_count=max_review_count,
            include_back_actions=False,
        )
    bucket_values = build_mistake_review_bucket_values(max_review_count)
    page_size = 6
    visible_values = bucket_values[:page_size]

    rows: list[list[InlineKeyboardButton]] = []
    current_row: list[InlineKeyboardButton] = []
    for value in visible_values:
        current_row.append(
            InlineKeyboardButton(
                "1-9" if value == 1 else str(value),
                callback_data=f"menu:mistakes_review_exact:{value}:0",
            )
        )
        if len(current_row) == 3:
            rows.append(current_row)
            current_row = []
    if current_row:
        rows.append(current_row)

    if len(bucket_values) > page_size:
        rows.append([InlineKeyboardButton("متابعة", callback_data="menu:mistakes_review_pick:1")])

    return InlineKeyboardMarkup(rows)


def build_mistake_bank_message(
    summary: dict[str, Any],
    *,
    has_open_session: bool,
) -> tuple[str, InlineKeyboardMarkup | None]:
    max_review_count = get_mistake_review_max_count(summary)
    return (
        build_mistake_bank_text(
            summary,
            max_review_count=max_review_count if int(summary["active"]) > 0 else None,
            has_open_session=has_open_session,
        ),
        build_mistake_bank_keyboard(
            max_review_count=max_review_count,
            has_open_session=has_open_session,
        ),
    )


def build_selector_label(position: int, status: str) -> str:
    if status == "correct":
        return f"✅{position}"
    if status == "wrong":
        return f"❌{position}"
    return f"◻️{position}"


def build_session_keyboard(
    session_payload: dict[str, Any],
    question_payload: dict[str, Any],
    states: list[dict[str, Any]],
    *,
    selector_mode: bool,
) -> InlineKeyboardMarkup:
    session_id = int(session_payload["session_id"])
    current_position = int(question_payload["position"])
    total_questions = int(session_payload["question_count"])

    if selector_mode:
        rows: list[list[InlineKeyboardButton]] = []
        current_row: list[InlineKeyboardButton] = []
        for item in states:
            current_row.append(
                InlineKeyboardButton(
                    build_selector_label(int(item["position"]), str(item["status"])),
                    callback_data=f"session:{session_id}:jump:{item['position']}",
                )
            )
            if len(current_row) == 4:
                rows.append(current_row)
                current_row = []
        if current_row:
            rows.append(current_row)
        rows.append(
            [
                InlineKeyboardButton("العودة للسؤال", callback_data=f"session:{session_id}:view"),
                InlineKeyboardButton("الملخص", callback_data=f"session:{session_id}:summary"),
            ]
        )
        return InlineKeyboardMarkup(rows)

    previous_callback = f"session:{session_id}:prev" if current_position > 1 else f"session:{session_id}:noop"
    next_label = "التالي"
    next_callback = f"session:{session_id}:next" if current_position < total_questions else f"session:{session_id}:finish"
    if current_position >= total_questions:
        next_label = "إنهاء"
    return InlineKeyboardMarkup(
        [
            [
                InlineKeyboardButton("السابق", callback_data=previous_callback),
                InlineKeyboardButton(f"{current_position}/{total_questions}", callback_data=f"session:{session_id}:noop"),
                InlineKeyboardButton(next_label, callback_data=next_callback),
            ],
            [
                InlineKeyboardButton("الخريطة", callback_data=f"session:{session_id}:overview"),
                InlineKeyboardButton("الملخص", callback_data=f"session:{session_id}:summary"),
            ],
        ]
    )


def build_session_caption(
    session_payload: dict[str, Any],
    question_payload: dict[str, Any],
    progress: dict[str, int],
    *,
    selector_mode: bool,
) -> str:
    session_title = "جلسة مراجعة الأخطاء" if session_payload["session_kind"] == "review" else "جلسة اليوم"
    if question_payload.get("selected_option") is None:
        answer_line = None
    elif int(question_payload.get("is_correct") or 0) == 1:
        answer_line = f"الحالة الحالية: إجابتك صحيحة ({OPTION_LABELS[str(question_payload['selected_option'])]})."
    else:
        answer_line = (
            f"الحالة الحالية: اخترت {OPTION_LABELS[str(question_payload['selected_option'])]} "
            f"والصحيحة {OPTION_LABELS[str(question_payload['correct_option'])]}."
        )

    lines = [
        f"{session_title} | السؤال {question_payload['position']} من {session_payload['question_count']}",
        f"التقدم: ✅ {progress['correct']} | ❌ {progress['wrong']} | ⏳ {progress['unanswered']}",
    ]
    if answer_line is not None:
        lines.append(answer_line)
    if not selector_mode and question_payload.get("selected_option") is not None:
        lines.append("هذا السؤال مجاب بالفعل، ويمكنك المتابعة أو الرجوع إلى الخريطة.")
    caption_text = str(question_payload.get("caption") or "").strip()
    if caption_text:
        lines.append("")
        lines.append(caption_text)
    if selector_mode:
        lines.append("")
        lines.append("الخريطة: ◻️ بدون إجابة | ✅ صحيحة | ❌ خاطئة")
    return "\n".join(lines)


def get_repository(application: Application) -> StudyRepository:
    return application.bot_data["repository"]


def get_settings(application: Application) -> Settings:
    return application.bot_data["settings"]


def ensure_private_chat(update: Update) -> bool:
    return bool(update.effective_chat and update.effective_chat.type == "private")


async def begin_plan_setup(update: Update, context: ContextTypes.DEFAULT_TYPE) -> None:
    user = update.effective_user
    if user is None:
        return

    repository = get_repository(context.application)
    existing_plan = repository.get_plan_by_telegram_id(user.id)
    selected_days = set(existing_plan["weekdays"] if existing_plan else [])
    reminder_time = existing_plan["reminder_time"] if existing_plan else None
    question_count = existing_plan["question_count"] if existing_plan else None
    review_weekday = existing_plan.get("review_weekday") if existing_plan else None

    context.user_data["plan_setup"] = {
        "days": selected_days,
        "reminder_time": reminder_time,
        "question_count": question_count,
        "review_weekday": review_weekday,
        "mode": "edit" if existing_plan is not None else "create",
        "step": "days",
    }

    if existing_plan is not None:
        plan_setup_state = context.user_data["plan_setup"]
        if update.callback_query is not None:
            await update.callback_query.answer()
            await update.effective_chat.send_message(
                build_plan_setup_summary(plan_setup_state),
                reply_markup=build_plan_edit_keyboard(plan_setup_state),
            )
            return

        if update.effective_message is not None:
            await update.effective_message.reply_text(
                build_plan_setup_summary(plan_setup_state),
                reply_markup=build_plan_edit_keyboard(plan_setup_state),
            )
        return

    if update.callback_query is not None:
        await update.callback_query.answer()
        await update.effective_chat.send_message(
            "اختر أيام الدراسة التي تناسبك. يمكنك اختيار أكثر من يوم، ثم اضغط تم.",
            reply_markup=build_days_keyboard(selected_days),
        )
        return

    if update.effective_message is not None:
        await update.effective_message.reply_text(
            "اختر أيام الدراسة التي تناسبك. يمكنك اختيار أكثر من يوم، ثم اضغط تم.",
            reply_markup=build_days_keyboard(selected_days),
        )


async def finish_plan_setup_for_user(
    *,
    repository: StudyRepository,
    settings: Settings,
    telegram_user: Any,
    state: dict[str, Any],
) -> dict[str, Any]:
    selected_days = sorted(state.get("days", set()))
    review_weekday = state.get("review_weekday")
    if len(selected_days) <= 1:
        review_weekday = None
    return repository.save_plan(
        telegram_user,
        weekdays=selected_days,
        reminder_time=str(state.get("reminder_time")),
        question_count=int(state.get("question_count")),
        review_weekday=review_weekday,
        timezone_name=settings.timezone_name,
    )


async def handle_plan_text_input(update: Update, context: ContextTypes.DEFAULT_TYPE) -> None:
    if not ensure_private_chat(update):
        return

    message = update.effective_message
    user = update.effective_user
    if message is None or user is None or not message.text:
        return

    state = context.user_data.get("plan_setup")
    repository = get_repository(context.application)
    settings = get_settings(context.application)

    if not state:
        plan = repository.get_plan_by_telegram_id(user.id)
        if plan is None or not plan["onboarding_completed"]:
            repository.upsert_telegram_user(user)
            await message.reply_text("سأبدأ معك الإعداد الآن بدون الحاجة إلى /start.")
            await begin_plan_setup(update, context)
        return

    step = state.get("step")
    mode = str(state.get("mode") or "create")

    if step in {"time_hour", "time_minute"}:
        await message.reply_text("اختر وقت التذكير من الأزرار الظاهرة فوق بدل كتابة الوقت يدويًا.")
        return

    if step == "count":
        question_count = parse_question_count_input(message.text)
        if question_count is None:
            await message.reply_text(
                f"أرسل رقمًا صحيحًا بين {MIN_QUESTION_COUNT} و {MAX_QUESTION_COUNT}.",
            )
            return
        state["question_count"] = question_count
        selected_days = sorted(state.get("days", set()))
        if mode == "edit":
            if len(selected_days) <= 1:
                state["review_weekday"] = None
            state["step"] = "menu"
            await message.reply_text(
                "تم تحديث عدد الأسئلة اليومي. يمكنك متابعة تعديل بقية العناصر أو حفظ الخطة الآن.",
            )
            await message.reply_text(
                build_plan_setup_summary(state),
                reply_markup=build_plan_edit_keyboard(state),
            )
            return
        if len(selected_days) <= 1:
            plan = await finish_plan_setup_for_user(
                repository=repository,
                settings=settings,
                telegram_user=user,
                state=state,
            )
            context.user_data.pop("plan_setup", None)
            await message.reply_text(
                "تم حفظ خطتك ✅\n\n"
                f"{build_plan_summary(plan)}\n\n"
                "من الآن فصاعدًا سأرسل لك جلسة خاصة بك في الأيام التي اخترتها.",
                reply_markup=build_user_menu_keyboard(context.application, plan),
            )
            return

        state["step"] = "review_day"
        await message.reply_text(
            "اختر يوم مراجعة الأخطاء من بين أيامك. في هذا اليوم سأعطيك نفس عددك اليومي لكن من بنك أخطائك."
            " إذا لم ترد يومًا مخصصًا، اختر بدون يوم مراجعة.",
            reply_markup=build_review_day_keyboard(selected_days, state.get("review_weekday")),
        )


async def mistakes_command(update: Update, context: ContextTypes.DEFAULT_TYPE) -> None:
    if not ensure_private_chat(update):
        return

    user = update.effective_user
    message = update.effective_message
    if user is None or message is None:
        return
    await send_mistake_bank_summary(context.application, user.id, reply_method=message.reply_text)


async def reset_command(update: Update, context: ContextTypes.DEFAULT_TYPE) -> None:
    if not ensure_private_chat(update):
        return

    user = update.effective_user
    message = update.effective_message
    if user is None or message is None:
        return

    repository = get_repository(context.application)
    was_reset = repository.reset_user(user.id)
    context.user_data.clear()

    if not was_reset:
        await message.reply_text(
            "لا توجد خطة محفوظة أو جلسات سابقة لهذا الحساب الآن. أرسل أي رسالة وسأبدأ الإعداد من جديد.",
        )
        return

    await message.reply_text(
        "تمت إعادة ضبط محادثتك وخطتك وسجل الأخطاء لهذا الحساب. أرسل أي رسالة وسأبدأ معك من الصفر.",
    )


async def start_command(update: Update, context: ContextTypes.DEFAULT_TYPE) -> None:
    if not ensure_private_chat(update):
        if update.effective_message is not None:
            await update.effective_message.reply_text("افتح البوت في الخاص حتى أجهز لك خطة شخصية.")
        return

    user = update.effective_user
    if user is None or update.effective_message is None:
        return

    repository = get_repository(context.application)
    repository.upsert_telegram_user(user)
    plan = repository.get_plan_by_telegram_id(user.id)
    start_payload = context.args[0] if context.args else ""

    if plan is None or not plan["onboarding_completed"]:
        welcome_text = "أهلا بك في عبقور. سأجهز لك خطة خفيفة تناسب أيامك، ولن أرسل لك إلا ما تحتاجه أنت فقط."
        if start_payload == CHANNEL_START_PAYLOAD:
            welcome_text = "أهلا بك. وصلت من القناة، ومن هنا سنكمل إعداد خطتك الشخصية الخاصة بك فقط."
        await update.effective_message.reply_text(welcome_text)
        await begin_plan_setup(update, context)
        return

    intro_prefix = ""
    if start_payload == CHANNEL_START_PAYLOAD:
        intro_prefix = "تم تحويلك من القناة إلى خطتك الخاصة.\n\n"

    await update.effective_message.reply_text(
        f"{intro_prefix}أهلا {plan['display_name']}\n\n{build_plan_summary(plan)}",
        reply_markup=build_user_menu_keyboard(context.application, plan),
    )


async def plan_command(update: Update, context: ContextTypes.DEFAULT_TYPE) -> None:
    if not ensure_private_chat(update):
        return
    await begin_plan_setup(update, context)


async def pause_command(update: Update, context: ContextTypes.DEFAULT_TYPE) -> None:
    user = update.effective_user
    if user is None or update.effective_message is None:
        return
    repository = get_repository(context.application)
    plan = repository.set_plan_active(user.id, False)
    if plan is None:
        await update.effective_message.reply_text("لم أجد خطة مفعلة بعد. ابدأ بالأمر /start.")
        return
    await update.effective_message.reply_text(
        "تم إيقاف التذكير. يمكنك تفعيله لاحقًا متى شئت.\n\n"
        f"{build_plan_summary(plan)}",
        reply_markup=build_user_menu_keyboard(context.application, plan),
    )


async def resume_command(update: Update, context: ContextTypes.DEFAULT_TYPE) -> None:
    user = update.effective_user
    if user is None or update.effective_message is None:
        return
    repository = get_repository(context.application)
    plan = repository.set_plan_active(user.id, True)
    if plan is None:
        await update.effective_message.reply_text("لم أجد خطة محفوظة بعد. ابدأ بالأمر /start.")
        return
    await update.effective_message.reply_text(
        "تم تفعيل التذكير من جديد ✅\n\n"
        f"{build_plan_summary(plan)}",
        reply_markup=build_user_menu_keyboard(context.application, plan),
    )


async def today_command(update: Update, context: ContextTypes.DEFAULT_TYPE) -> None:
    if not ensure_private_chat(update):
        return
    user = update.effective_user
    if user is None:
        return
    await start_or_resume_today_session(context.application, user.id, manual_trigger=True)


async def send_mistake_bank_summary(application: Application, telegram_user_id: int, *, reply_method: Any) -> None:
    repository = get_repository(application)
    summary = repository.get_mistake_bank_summary(telegram_user_id)
    if summary is None:
        await reply_method("ابدأ أولًا عبر /start حتى أجهز لك خطة شخصية.")
        return

    open_review_session = repository.get_latest_open_manual_review_session(telegram_user_id)
    text, reply_markup = build_mistake_bank_message(
        summary,
        has_open_session=open_review_session is not None,
    )
    await reply_method(
        text,
        reply_markup=reply_markup,
    )


async def edit_mistake_bank_summary_message(
    application: Application,
    telegram_user_id: int,
    *,
    query: Any,
) -> None:
    repository = get_repository(application)
    summary = repository.get_mistake_bank_summary(telegram_user_id)
    if summary is None:
        if query.message is not None:
            await query.edit_message_text("ابدأ أولًا عبر /start حتى أجهز لك خطة شخصية.")
        else:
            await application.bot.send_message(
                chat_id=telegram_user_id,
                text="ابدأ أولًا عبر /start حتى أجهز لك خطة شخصية.",
            )
        return

    open_review_session = repository.get_latest_open_manual_review_session(telegram_user_id)
    text, reply_markup = build_mistake_bank_message(
        summary,
        has_open_session=open_review_session is not None,
    )
    if query.message is not None:
        await query.edit_message_text(text, reply_markup=reply_markup)
    else:
        await application.bot.send_message(
            chat_id=telegram_user_id,
            text=text,
            reply_markup=reply_markup,
        )


async def show_mistake_review_count_picker(
    application: Application,
    telegram_user_id: int,
    *,
    query: Any,
    page: int,
) -> None:
    repository = get_repository(application)
    summary = repository.get_mistake_bank_summary(telegram_user_id)
    if summary is None or int(summary["active"]) <= 0:
        await edit_mistake_bank_summary_message(application, telegram_user_id, query=query)
        return

    max_review_count = get_mistake_review_max_count(summary)
    if max_review_count < 10:
        text = build_mistake_review_exact_text(bucket_start=1, max_review_count=max_review_count)
        reply_markup = build_mistake_review_exact_keyboard(
            bucket_start=1,
            bucket_page=0,
            max_review_count=max_review_count,
        )
    else:
        text = build_mistake_review_bucket_text(max_review_count=max_review_count, page=page)
        reply_markup = build_mistake_review_bucket_keyboard(max_review_count=max_review_count, page=page)

    if query.message is not None:
        await query.edit_message_text(text, reply_markup=reply_markup)
    else:
        await application.bot.send_message(
            chat_id=telegram_user_id,
            text=text,
            reply_markup=reply_markup,
        )


async def show_mistake_review_exact_count_picker(
    application: Application,
    telegram_user_id: int,
    *,
    query: Any,
    bucket_start: int,
    bucket_page: int,
) -> None:
    repository = get_repository(application)
    summary = repository.get_mistake_bank_summary(telegram_user_id)
    if summary is None or int(summary["active"]) <= 0:
        await edit_mistake_bank_summary_message(application, telegram_user_id, query=query)
        return

    max_review_count = get_mistake_review_max_count(summary)
    if max_review_count < 10:
        bucket_start = 1
    exact_values = build_mistake_review_exact_values(bucket_start, max_review_count)
    if not exact_values:
        await show_mistake_review_count_picker(
            application,
            telegram_user_id,
            query=query,
            page=bucket_page,
        )
        return

    text = build_mistake_review_exact_text(bucket_start=bucket_start, max_review_count=max_review_count)
    reply_markup = build_mistake_review_exact_keyboard(
        bucket_start=bucket_start,
        bucket_page=bucket_page,
        max_review_count=max_review_count,
    )
    if query.message is not None:
        await query.edit_message_text(text, reply_markup=reply_markup)
    else:
        await application.bot.send_message(
            chat_id=telegram_user_id,
            text=text,
            reply_markup=reply_markup,
        )


async def start_or_resume_mistake_review_session(
    application: Application,
    telegram_user_id: int,
    *,
    manual_trigger: bool,
    question_limit: int | None = None,
    source_message_id: int | None = None,
    source_query: Any | None = None,
) -> None:
    repository = get_repository(application)
    settings = get_settings(application)

    plan = repository.get_plan_by_telegram_id(telegram_user_id)
    if plan is None:
        if source_query is not None and source_query.message is not None:
            await source_query.edit_message_text("ابدأ أولًا عبر /start حتى أجهز لك خطة شخصية.")
        else:
            await application.bot.send_message(
                chat_id=telegram_user_id,
                text="ابدأ أولًا عبر /start حتى أجهز لك خطة شخصية.",
            )
        return

    existing_session = repository.get_latest_open_manual_review_session(telegram_user_id)
    if existing_session is not None:
        if manual_trigger:
            await cleanup_session_messages(application, existing_session)
            if source_message_id is not None:
                await delete_message_safely(application, telegram_user_id, source_message_id)
            notice_message = await application.bot.send_message(
                chat_id=telegram_user_id,
                text="نكمل مراجعة الأخطاء من حيث توقفت 👌",
            )
            repository.set_session_notice_message_id(existing_session["session_id"], int(notice_message.message_id))
        await render_session_view(application, existing_session["session_id"], chat_id=telegram_user_id)
        return

    current_date = datetime.now(settings.timezone).date()
    session = repository.create_manual_review_session(
        telegram_user_id,
        current_date,
        question_limit=question_limit,
    )
    if session is None:
        if source_query is not None:
            await edit_mistake_bank_summary_message(application, telegram_user_id, query=source_query)
        else:
            await application.bot.send_message(
                chat_id=telegram_user_id,
                text="لا توجد أسئلة نشطة في بنك الأخطاء الآن. عندما تخطئ في سؤال سأضيفه هنا لتراجعه لاحقًا.",
            )
        return

    if source_message_id is not None:
        await delete_message_safely(application, telegram_user_id, source_message_id)

    notice_message = await application.bot.send_message(
        chat_id=telegram_user_id,
        text=(
            f"مراجعة الأخطاء جاهزة ✨\n"
            f"عدد الأسئلة في هذه الجلسة: {session['question_count']}\n"
            "هذه جلسة اختيارية من بنك الأخطاء، ويمكنك التنقل بين الأسئلة وفتح الخريطة في أي وقت."
        ),
    )
    repository.set_session_notice_message_id(session["session_id"], int(notice_message.message_id))
    repository.mark_session_delivered(session["session_id"])
    await render_session_view(application, session["session_id"], chat_id=telegram_user_id)


async def handle_private_chat_membership(update: Update, context: ContextTypes.DEFAULT_TYPE) -> None:
    membership_update = update.my_chat_member
    if membership_update is None:
        return

    chat = membership_update.chat
    user = membership_update.from_user
    if chat.type != "private" or user is None:
        return

    repository = get_repository(context.application)
    repository.upsert_telegram_user(user)

    new_status = str(membership_update.new_chat_member.status)
    if new_status in {"kicked", "left"}:
        repository.set_plan_active(user.id, False)


async def handle_menu_callback(update: Update, context: ContextTypes.DEFAULT_TYPE) -> None:
    query = update.callback_query
    if query is None or query.data is None:
        return
    user = query.from_user
    action = query.data.split(":", maxsplit=1)[1]
    action_parts = action.split(":")
    action_name = action_parts[0]

    if action == "plan":
        if query.message is not None:
            await delete_message_safely(context.application, int(user.id), int(query.message.message_id))
        await begin_plan_setup(update, context)
        return

    await query.answer()
    if action_name == "today":
        if query.message is not None:
            await delete_message_safely(context.application, int(user.id), int(query.message.message_id))
        await start_or_resume_today_session(context.application, user.id, manual_trigger=True)
        return

    if action == "mistakes_review":
        await start_or_resume_mistake_review_session(
            context.application,
            user.id,
            manual_trigger=True,
            source_message_id=int(query.message.message_id) if query.message is not None else None,
            source_query=query,
        )
        return

    if action_name == "mistakes_review_pick":
        page = 0
        if len(action_parts) >= 2:
            try:
                page = max(0, int(action_parts[1]))
            except ValueError:
                page = 0
        await show_mistake_review_count_picker(
            context.application,
            user.id,
            query=query,
            page=page,
        )
        return

    if action_name == "mistakes_review_exact" and len(action_parts) >= 3:
        try:
            bucket_start = int(action_parts[1])
            bucket_page = max(0, int(action_parts[2]))
        except ValueError:
            return
        await show_mistake_review_exact_count_picker(
            context.application,
            user.id,
            query=query,
            bucket_start=bucket_start,
            bucket_page=bucket_page,
        )
        return

    if action_name == "mistakes_review_start" and len(action_parts) >= 2:
        try:
            question_limit = int(action_parts[1])
        except ValueError:
            return
        await start_or_resume_mistake_review_session(
            context.application,
            user.id,
            manual_trigger=True,
            question_limit=question_limit,
            source_message_id=int(query.message.message_id) if query.message is not None else None,
            source_query=query,
        )
        return

    repository = get_repository(context.application)
    if action_name == "mistakes":
        await edit_mistake_bank_summary_message(context.application, user.id, query=query)
        return

    if action_name == "pause":
        plan = repository.set_plan_active(user.id, False)
        if plan is not None:
            if query.message is not None:
                await delete_message_safely(context.application, int(user.id), int(query.message.message_id))
            await context.application.bot.send_message(
                chat_id=int(user.id),
                text=(
                    "تم إيقاف التذكير. جلساتك لن تصلك تلقائيًا حتى تعيد تفعيله.\n\n"
                    f"{build_plan_summary(plan)}"
                ),
                reply_markup=build_user_menu_keyboard(context.application, plan),
            )
        return

    if action == "resume":
        plan = repository.set_plan_active(user.id, True)
        if plan is not None:
            if query.message is not None:
                await delete_message_safely(context.application, int(user.id), int(query.message.message_id))
            await context.application.bot.send_message(
                chat_id=int(user.id),
                text=(
                    "تم تفعيل التذكير ✅\n\n"
                    f"{build_plan_summary(plan)}"
                ),
                reply_markup=build_user_menu_keyboard(context.application, plan),
            )


async def handle_setup_callback(update: Update, context: ContextTypes.DEFAULT_TYPE) -> None:
    query = update.callback_query
    if query is None or query.data is None or query.from_user is None:
        return

    repository = get_repository(context.application)
    settings = get_settings(context.application)
    state = context.user_data.setdefault(
        "plan_setup",
        {"days": set(), "reminder_time": None, "question_count": None, "review_weekday": None, "step": "days"},
    )
    parts = query.data.split(":")
    action = parts[1]
    mode = str(state.get("mode") or "create")

    if action == "cancel":
        await query.answer()
        context.user_data.pop("plan_setup", None)
        plan = repository.get_plan_by_telegram_id(query.from_user.id)
        if plan is not None:
            await query.edit_message_text(
                "تم إلغاء تعديل الخطة.\n\n"
                f"{build_plan_summary(plan)}",
                reply_markup=build_user_menu_keyboard(context.application, plan),
            )
        else:
            await query.edit_message_text("تم إلغاء الإعداد. أرسل /start عندما تريد البدء.")
        return

    if action == "save":
        selected_days = sorted(state.get("days", set()))
        if not selected_days:
            await query.answer("اختر يومًا واحدًا على الأقل.", show_alert=True)
            return
        if not state.get("reminder_time") or state.get("question_count") is None:
            await query.answer("أكمل وقت التذكير وعدد الأسئلة أولًا.", show_alert=True)
            return
        plan = await finish_plan_setup_for_user(
            repository=repository,
            settings=settings,
            telegram_user=query.from_user,
            state=state,
        )
        context.user_data.pop("plan_setup", None)
        await query.answer()
        await query.edit_message_text(
            "تم حفظ خطتك ✅\n\n"
            f"{build_plan_summary(plan)}\n\n"
            "من الآن فصاعدًا سأعمل بالخطة الجديدة.",
            reply_markup=build_user_menu_keyboard(context.application, plan),
        )
        return

    if action == "edit" and len(parts) >= 3:
        edit_target = parts[2]
        if edit_target == "days":
            state["step"] = "days"
            await query.answer()
            await query.edit_message_text(
                "اختر أيام الدراسة التي تناسبك. يمكنك اختيار أكثر من يوم، ثم اضغط تم.",
                reply_markup=build_days_keyboard(state.setdefault("days", set())),
            )
            return
        if edit_target == "time":
            reminder_time = state.get("reminder_time")
            selected_hour = None
            if isinstance(reminder_time, str) and ":" in reminder_time:
                hour_raw, _, _ = reminder_time.partition(":")
                if hour_raw.isdigit():
                    selected_hour = int(hour_raw)
            state["step"] = "time_hour"
            await query.answer()
            await query.edit_message_text(
                "اختر الآن ساعة التذكير اليومية.",
                reply_markup=build_time_hour_keyboard(selected_hour),
            )
            return
        if edit_target == "count":
            state["step"] = "count"
            await query.answer()
            await query.edit_message_text(
                "اختر عدد الأسئلة اليومي من الخيارات السريعة، أو أرسل رقمًا يدويًا في المحادثة.",
                reply_markup=build_question_count_keyboard(state.get("question_count")),
            )
            return
        if edit_target == "review":
            selected_days = sorted(state.get("days", set()))
            if len(selected_days) <= 1:
                await query.answer("يوم مراجعة الأخطاء يحتاج يومي دراسة على الأقل.", show_alert=True)
                return
            state["step"] = "review_day"
            await query.answer()
            await query.edit_message_text(
                "اختر يوم مراجعة الأخطاء من بين أيامك. في هذا اليوم سأعطيك أسئلة من بنك أخطائك.",
                reply_markup=build_review_day_keyboard(selected_days, state.get("review_weekday")),
            )
            return

    if action == "day":
        await query.answer()
        weekday = int(parts[2])
        selected_days: set[int] = state.setdefault("days", set())
        if weekday in selected_days:
            selected_days.remove(weekday)
        else:
            selected_days.add(weekday)
        await query.edit_message_reply_markup(reply_markup=build_days_keyboard(selected_days))
        return

    if action == "days" and parts[2] == "done":
        selected_days = state.setdefault("days", set())
        if not selected_days:
            await query.answer("اختر يومًا واحدًا على الأقل.", show_alert=True)
            return
        if len(selected_days) <= 1:
            state["review_weekday"] = None
        if mode == "edit":
            state["step"] = "menu"
            await query.answer()
            await query.edit_message_text(
                build_plan_setup_summary(state),
                reply_markup=build_plan_edit_keyboard(state),
            )
            return
        selected_hour = None
        reminder_time = state.get("reminder_time")
        if isinstance(reminder_time, str) and ":" in reminder_time:
            hour_raw, _, _ = reminder_time.partition(":")
            if hour_raw.isdigit():
                selected_hour = int(hour_raw)
        state["step"] = "time_hour"
        await query.answer()
        await query.edit_message_text(
            f"تم اختيار الأيام: {format_weekdays(sorted(selected_days))}\n\n"
            "اختر الآن ساعة التذكير اليومية.",
            reply_markup=build_time_hour_keyboard(selected_hour),
        )
        return

    if action == "timeback":
        reminder_time = state.get("reminder_time")
        selected_hour = None
        if isinstance(reminder_time, str) and ":" in reminder_time:
            hour_raw, _, _ = reminder_time.partition(":")
            if hour_raw.isdigit():
                selected_hour = int(hour_raw)
        state["step"] = "time_hour"
        state.pop("selected_hour", None)
        await query.answer()
        await query.edit_message_text(
            "اختر ساعة التذكير اليومية.",
            reply_markup=build_time_hour_keyboard(selected_hour),
        )
        return

    if action == "timehour" and len(parts) >= 3:
        try:
            selected_hour = int(parts[2])
        except ValueError:
            return
        state["selected_hour"] = selected_hour
        state["step"] = "time_minute"
        await query.answer()
        await query.edit_message_text(
            f"تم اختيار الساعة: {format_time_label(f'{selected_hour:02d}:00')}\n\nاختر الآن الدقائق.",
            reply_markup=build_time_minute_keyboard(selected_hour),
        )
        return

    if action == "timeminute" and len(parts) >= 3:
        try:
            selected_minute = int(parts[2])
        except ValueError:
            return
        selected_hour = state.get("selected_hour")
        if not isinstance(selected_hour, int):
            state["step"] = "time_hour"
            await query.answer("اختر الساعة أولًا.", show_alert=True)
            await query.edit_message_text(
                "اختر ساعة التذكير اليومية.",
                reply_markup=build_time_hour_keyboard(),
            )
            return
        reminder_time = f"{selected_hour:02d}:{selected_minute:02d}"
        state["reminder_time"] = reminder_time
        if mode == "edit":
            state["step"] = "menu"
            await query.answer()
            await query.edit_message_text(
                build_plan_setup_summary(state),
                reply_markup=build_plan_edit_keyboard(state),
            )
            return
        state["step"] = "count"
        await query.answer()
        await query.edit_message_text(
            f"تم حفظ الوقت: {format_time_label(reminder_time)}\n\nاختر الآن عدد الأسئلة اليومي أو أرسل رقمًا يدويًا.",
            reply_markup=build_question_count_keyboard(state.get("question_count")),
        )
        return

    if action == "count" and len(parts) >= 3:
        try:
            question_count = int(parts[2])
        except ValueError:
            return
        if question_count < MIN_QUESTION_COUNT or question_count > MAX_QUESTION_COUNT:
            return
        state["question_count"] = question_count
        selected_days = sorted(state.get("days", set()))
        if mode == "edit":
            if len(selected_days) <= 1:
                state["review_weekday"] = None
            state["step"] = "menu"
            await query.answer()
            await query.edit_message_text(
                build_plan_setup_summary(state),
                reply_markup=build_plan_edit_keyboard(state),
            )
            return
        if len(selected_days) <= 1:
            plan = await finish_plan_setup_for_user(
                repository=repository,
                settings=settings,
                telegram_user=query.from_user,
                state=state,
            )
            context.user_data.pop("plan_setup", None)
            await query.answer()
            await query.edit_message_text(
                "تم حفظ خطتك ✅\n\n"
                f"{build_plan_summary(plan)}\n\n"
                "من الآن فصاعدًا سأرسل لك جلسة خاصة بك في الأيام التي اخترتها.",
                reply_markup=build_user_menu_keyboard(context.application, plan),
            )
            return
        state["step"] = "review_day"
        await query.answer()
        await query.edit_message_text(
            "اختر يوم مراجعة الأخطاء من بين أيامك. في هذا اليوم سأعطيك نفس عددك اليومي لكن من بنك أخطائك."
            " إذا لم ترد يومًا مخصصًا، اختر بدون يوم مراجعة.",
            reply_markup=build_review_day_keyboard(selected_days, state.get("review_weekday")),
        )
        return

    if action == "review":
        await query.answer()
        review_value = parts[2]
        state["review_weekday"] = None if review_value == "none" else int(review_value)
        if mode == "edit":
            state["step"] = "menu"
            await query.edit_message_text(
                build_plan_setup_summary(state),
                reply_markup=build_plan_edit_keyboard(state),
            )
            return
        plan = await finish_plan_setup_for_user(
            repository=repository,
            settings=settings,
            telegram_user=query.from_user,
            state=state,
        )
        context.user_data.pop("plan_setup", None)
        await query.edit_message_text(
            "تم حفظ خطتك ✅\n\n"
            f"{build_plan_summary(plan)}\n\n"
            "من الآن فصاعدًا سأرسل لك جلسة خاصة بك في الأيام التي اخترتها.",
            reply_markup=build_user_menu_keyboard(context.application, plan),
        )
        return


async def complete_session(application: Application, session_id: int, telegram_user_id: int) -> None:
    repository = get_repository(application)
    repository.mark_session_completed(session_id)
    session = repository.get_session_by_id(session_id)
    if session is not None:
        await cleanup_session_messages(application, session)
    result = repository.get_session_result(session_id)
    plan = repository.get_plan_by_telegram_id(telegram_user_id)
    session_title = "مراجعة الأخطاء" if session is not None and session["session_kind"] == "review" else "جلسة اليوم"
    extra_lines: list[str] = []
    if session is not None and session["session_kind"] == "review":
        summary = repository.get_mistake_bank_summary(telegram_user_id)
        active_remaining = int(summary["active"]) if summary is not None else 0
        if active_remaining == 0:
            extra_lines.append("بنك الأخطاء أصبح فارغًا الآن. كل ما راجعته بشكل صحيح خرج منه.")
        else:
            extra_lines.append(f"المتبقي الآن في بنك الأخطاء: {active_remaining} سؤال.")
    await application.bot.send_message(
        chat_id=telegram_user_id,
        text=(
            f"لقد أنهيت {session_title} ✨\n"
            f"الإجابات الصحيحة: {result['correct']} من {result['total']}\n"
            + ("\n".join(extra_lines) + "\n" if extra_lines else "")
            ).rstrip(),
        reply_markup=build_user_menu_keyboard(application, plan) if plan is not None else None,
    )


async def delete_message_safely(application: Application, chat_id: int, message_id: int | None) -> None:
    if message_id is None:
        return
    try:
        await application.bot.delete_message(chat_id=chat_id, message_id=int(message_id))
    except TelegramError:
        pass


async def clear_active_poll_message(application: Application, session_payload: dict[str, Any]) -> None:
    repository = get_repository(application)
    await delete_message_safely(
        application,
        int(session_payload["telegram_user_id"]),
        session_payload.get("active_poll_message_id"),
    )
    repository.clear_active_poll(int(session_payload["session_id"]))


async def cleanup_session_messages(application: Application, session_payload: dict[str, Any]) -> None:
    repository = get_repository(application)
    await clear_active_poll_message(application, session_payload)
    await delete_message_safely(
        application,
        int(session_payload["telegram_user_id"]),
        session_payload.get("notice_message_id"),
    )
    repository.set_session_notice_message_id(int(session_payload["session_id"]), None)
    await delete_message_safely(
        application,
        int(session_payload["telegram_user_id"]),
        session_payload.get("view_message_id"),
    )
    repository.set_session_view_message_id(int(session_payload["session_id"]), None)


async def send_question_poll(
    application: Application,
    session_payload: dict[str, Any],
    question_payload: dict[str, Any],
) -> None:
    repository = get_repository(application)
    await clear_active_poll_message(application, session_payload)

    if question_payload.get("selected_option") is not None:
        return

    poll_message = await application.bot.send_poll(
        chat_id=int(session_payload["telegram_user_id"]),
        question=f"اختر إجابة السؤال {question_payload['position']} من {session_payload['question_count']}",
        options=[OPTION_LABELS[option] for option in POLL_OPTION_ORDER],
        type="quiz",
        is_anonymous=False,
        correct_option_id=POLL_OPTION_ORDER.index(str(question_payload["correct_option"])),
    )
    if poll_message.poll is None:
        return

    repository.set_active_poll(
        int(session_payload["session_id"]),
        poll_id=poll_message.poll.id,
        poll_message_id=int(poll_message.message_id),
        session_question_id=int(question_payload["session_question_id"]),
    )


async def render_session_view(
    application: Application,
    session_id: int,
    *,
    chat_id: int | None = None,
    selector_mode: bool = False,
    refresh_poll: bool = True,
) -> None:
    repository = get_repository(application)
    session = repository.get_session_by_id(session_id)
    if session is None:
        return

    question_payload = repository.get_session_question_by_position(
        session_id,
        int(session["current_position"]),
    )
    if question_payload is None:
        return

    repository.mark_session_started(session_id)
    progress = repository.get_session_progress(session_id)
    caption = build_session_caption(session, question_payload, progress, selector_mode=selector_mode)
    reply_markup = build_session_keyboard(session, question_payload, repository.list_session_question_states(session_id), selector_mode=selector_mode)
    target_chat_id = int(chat_id or session["telegram_user_id"])

    if selector_mode:
        await clear_active_poll_message(application, session)

    await delete_message_safely(application, target_chat_id, session.get("view_message_id"))
    repository.set_session_view_message_id(session_id, None)

    with Path(str(question_payload["image_path"])).open("rb") as image_file:
        sent_message = await application.bot.send_photo(
            chat_id=target_chat_id,
            photo=image_file,
            caption=caption,
            reply_markup=reply_markup,
        )
    repository.set_session_view_message_id(session_id, int(sent_message.message_id))

    if selector_mode or not refresh_poll:
        return

    refreshed_session = repository.get_session_by_id(session_id)
    if refreshed_session is None:
        return
    await send_question_poll(application, refreshed_session, question_payload)


async def start_or_resume_today_session(
    application: Application,
    telegram_user_id: int,
    *,
    manual_trigger: bool,
) -> None:
    repository = get_repository(application)
    settings = get_settings(application)
    current_date = datetime.now(settings.timezone).date()

    plan = repository.get_plan_by_telegram_id(telegram_user_id)
    if plan is None:
        await application.bot.send_message(
            chat_id=telegram_user_id,
            text="ابدأ أولًا عبر /start حتى أجهز لك خطة شخصية.",
        )
        return

    existing_session = repository.get_latest_open_session_for_user(telegram_user_id)
    if existing_session is not None and existing_session["scheduled_for"] == current_date.isoformat():
        if manual_trigger and existing_session.get("started_at") is not None:
            await cleanup_session_messages(application, existing_session)
            notice_message = await application.bot.send_message(chat_id=telegram_user_id, text="نكمل من حيث توقفت 👌")
            repository.set_session_notice_message_id(existing_session["session_id"], int(notice_message.message_id))
        await render_session_view(application, existing_session["session_id"], chat_id=telegram_user_id)
        return

    session = repository.get_session_for_user_and_date(telegram_user_id, current_date)
    if session is None:
        session = repository.create_session_for_user(telegram_user_id, current_date)
    elif session["status"] == "completed" and manual_trigger:
        session = repository.reset_completed_session_for_date(telegram_user_id, current_date)

    if session is None:
        await application.bot.send_message(
            chat_id=telegram_user_id,
            text="لا توجد أسئلة كافية في بنك الأسئلة الآن. سأرسل لك جلسة جديدة حالما تجهز أسئلة إضافية.",
        )
        return

    if session["status"] == "completed":
        await complete_session(application, session["session_id"], telegram_user_id)
        return

    if session["delivered_at"] is None:
        session_title = "مراجعة الأخطاء" if session["session_kind"] == "review" else "جلسة اليوم"
        intro_text = (
            f"{session_title} جاهزة ✨\n"
            f"عدد الأسئلة اليوم: {session['question_count']}\n"
            "ستصلك صورة السؤال ومعها استطلاع للإجابة، ويمكنك التنقل بين الأسئلة وفتح الخريطة في أي وقت."
        )
        notice_message = await application.bot.send_message(chat_id=telegram_user_id, text=intro_text)
        repository.set_session_notice_message_id(session["session_id"], int(notice_message.message_id))
        repository.mark_session_delivered(session["session_id"])
    elif manual_trigger:
        notice_message = await application.bot.send_message(chat_id=telegram_user_id, text="نكمل من حيث توقفت 👌")
        repository.set_session_notice_message_id(session["session_id"], int(notice_message.message_id))

    await render_session_view(application, session["session_id"], chat_id=telegram_user_id)


async def schedule_due_sessions(application: Application) -> None:
    repository = get_repository(application)
    settings = get_settings(application)
    current_time = datetime.now(settings.timezone)
    due_plans = repository.list_due_plans(current_time)

    for plan in due_plans:
        session = repository.create_session_for_user(plan["telegram_user_id"], current_time.date())
        if session is None:
            continue
        if session["delivered_at"] is not None:
            continue

        try:
            session_title = "مراجعة الأخطاء" if session["session_kind"] == "review" else "جلسة اليوم"
            notice_message = await application.bot.send_message(
                chat_id=plan["telegram_user_id"],
                text=(
                    f"{session_title} جاهزة ✨\n"
                    f"عدد الأسئلة اليوم: {session['question_count']}"
                ),
                reply_markup=build_session_start_keyboard(),
            )
            repository.set_session_notice_message_id(session["session_id"], int(notice_message.message_id))
            repository.mark_session_delivered(session["session_id"])
        except Forbidden:
            repository.set_plan_active(plan["telegram_user_id"], False)
        except TelegramError:
            continue


async def scheduler_loop(application: Application) -> None:
    while True:
        try:
            await schedule_due_sessions(application)
        except asyncio.CancelledError:
            raise
        except Exception:
            pass
        await asyncio.sleep(SCHEDULER_INTERVAL_SECONDS)


async def handle_poll_answer(update: Update, context: ContextTypes.DEFAULT_TYPE) -> None:
    poll_answer = update.poll_answer
    if poll_answer is None or not poll_answer.option_ids:
        return

    repository = get_repository(context.application)
    poll_context = repository.get_active_poll_context(poll_answer.poll_id)
    if poll_context is None:
        return

    if int(poll_context["telegram_user_id"]) != int(poll_answer.user.id):
        return

    option_index = int(poll_answer.option_ids[0])
    if option_index < 0 or option_index >= len(POLL_OPTION_ORDER):
        return
    selected_option = POLL_OPTION_ORDER[option_index]

    answer_record = repository.record_answer(int(poll_context["session_question_id"]), selected_option)
    if answer_record is None:
        return
    if bool(answer_record.get("already_answered")):
        return

    repository.set_session_current_position(int(answer_record["session_id"]), int(answer_record["position"]))

    if int(answer_record["remaining"]) == 0:
        return


async def handle_session_navigation(update: Update, context: ContextTypes.DEFAULT_TYPE) -> None:
    query = update.callback_query
    if query is None or query.data is None or query.from_user is None:
        return

    parts = query.data.split(":")
    if len(parts) < 3:
        return

    try:
        session_id = int(parts[1])
    except ValueError:
        return

    repository = get_repository(context.application)
    session = repository.get_session_by_id(session_id)
    if session is None:
        await query.answer("لم أعد أجد هذه الجلسة.", show_alert=True)
        return
    if int(session["telegram_user_id"]) != int(query.from_user.id):
        await query.answer("هذه الجلسة ليست لك.", show_alert=True)
        return

    action = parts[2]
    if action == "noop":
        await query.answer()
        return

    if action == "summary":
        progress = repository.get_session_progress(session_id)
        await query.answer(
            f"✅ {progress['correct']} | ❌ {progress['wrong']} | ⏳ {progress['unanswered']}",
            show_alert=True,
        )
        return

    if action == "finish":
        progress = repository.get_session_progress(session_id)
        current_question = repository.get_session_question_by_position(session_id, int(session["current_position"]))
        if current_question is not None and current_question.get("selected_option") is None:
            await query.answer("أجب عن هذا السؤال أولًا، ثم أنهِ الجلسة متى أردت.", show_alert=True)
            return
        if int(progress["unanswered"]) > 0:
            await query.answer("أكمل بقية الأسئلة أولًا، أو ارجع لها من الخريطة.", show_alert=True)
            return
        await query.answer()
        await complete_session(
            context.application,
            session_id,
            int(session["telegram_user_id"]),
        )
        return

    if action == "prev":
        repository.set_session_current_position(session_id, int(session["current_position"]) - 1)
        await query.answer()
        await render_session_view(context.application, session_id)
        return

    if action == "next":
        repository.set_session_current_position(session_id, int(session["current_position"]) + 1)
        await query.answer()
        await render_session_view(context.application, session_id)
        return

    if action == "jump" and len(parts) >= 4:
        try:
            target_position = int(parts[3])
        except ValueError:
            return
        repository.set_session_current_position(session_id, target_position)
        await query.answer()
        await render_session_view(context.application, session_id)
        return

    if action == "overview":
        await query.answer()
        await render_session_view(context.application, session_id, selector_mode=True)
        return

    if action == "view":
        await query.answer()
        await render_session_view(context.application, session_id)


def build_application(settings: Settings | None = None) -> FastAPI:
    active_settings = settings or load_settings()
    active_settings.data_dir.mkdir(parents=True, exist_ok=True)
    active_settings.uploads_dir.mkdir(parents=True, exist_ok=True)

    repository = StudyRepository(
        active_settings.database_path,
        active_settings.uploads_dir,
        active_settings.legacy_question_file,
    )
    telegram_app = ApplicationBuilder().token(active_settings.telegram_bot_token).build()

    telegram_app.add_handler(CommandHandler("start", start_command))
    telegram_app.add_handler(CommandHandler("plan", plan_command))
    telegram_app.add_handler(CommandHandler("today", today_command))
    telegram_app.add_handler(CommandHandler("mistakes", mistakes_command))
    telegram_app.add_handler(CommandHandler("reset", reset_command))
    telegram_app.add_handler(CommandHandler("pause", pause_command))
    telegram_app.add_handler(CommandHandler("resume", resume_command))
    telegram_app.add_handler(CallbackQueryHandler(handle_setup_callback, pattern=r"^setup:"))
    telegram_app.add_handler(CallbackQueryHandler(handle_menu_callback, pattern=r"^menu:"))
    telegram_app.add_handler(CallbackQueryHandler(handle_session_navigation, pattern=r"^session:"))
    telegram_app.add_handler(PollAnswerHandler(handle_poll_answer))
    telegram_app.add_handler(ChatMemberHandler(handle_private_chat_membership, ChatMemberHandler.MY_CHAT_MEMBER))
    telegram_app.add_handler(MessageHandler(filters.TEXT & ~filters.COMMAND, handle_plan_text_input))

    telegram_app.bot_data["repository"] = repository
    telegram_app.bot_data["settings"] = active_settings

    app = FastAPI(title="Abqoor Study Bot")
    app.state.settings = active_settings
    app.state.repository = repository
    app.state.telegram_app = telegram_app
    app.state.scheduler_task = None

    app.mount("/static", StaticFiles(directory=STATIC_DIR), name="static")
    app.mount("/uploads", StaticFiles(directory=active_settings.uploads_dir), name="uploads")

    @app.on_event("startup")
    async def startup_event() -> None:
        repository.initialize()
        app.state.bot_username = ""

        if not active_settings.telegram_enabled:
            return

        await telegram_app.initialize()
        await telegram_app.start()
        bot_profile = await telegram_app.bot.get_me()
        app.state.bot_username = bot_profile.username or ""

        if telegram_app.updater is None:
            raise RuntimeError("تعذر تشغيل محدث تيليجرام.")

        await telegram_app.updater.start_polling(allowed_updates=Update.ALL_TYPES)
        app.state.scheduler_task = asyncio.create_task(scheduler_loop(telegram_app))

    @app.on_event("shutdown")
    async def shutdown_event() -> None:
        scheduler_task: asyncio.Task[Any] | None = app.state.scheduler_task
        if scheduler_task is not None:
            scheduler_task.cancel()
            try:
                await scheduler_task
            except asyncio.CancelledError:
                pass

        if not active_settings.telegram_enabled:
            repository.close()
            return

        if telegram_app.updater is not None:
            await telegram_app.updater.stop()
        await telegram_app.stop()
        await telegram_app.shutdown()
        repository.close()

    @app.get("/")
    async def read_index() -> FileResponse:
        return FileResponse(STATIC_DIR / "index.html")

    @app.get("/healthz")
    async def health_check() -> dict[str, Any]:
        return {
            "ok": True,
            "database_exists": repository.database_path.exists(),
            "uploads_dir_exists": active_settings.uploads_dir.exists(),
            "telegram_enabled": active_settings.telegram_enabled,
            "timezone_name": active_settings.timezone_name,
        }

    @app.post("/api/login")
    async def login(payload: dict[str, str]) -> dict[str, bool]:
        password = payload.get("password", "")
        validate_admin_token(active_settings, password)
        return {"ok": True}

    @app.get("/api/config")
    async def get_config(admin_token: str | None = Header(default=None, alias="X-Admin-Token")) -> dict[str, Any]:
        validate_admin_token(active_settings, admin_token)
        stats = repository.get_dashboard_stats()
        bot_username = getattr(app.state, "bot_username", "")
        return {
            "bot_username": bot_username,
            "bot_deep_link": build_bot_deep_link(bot_username),
            "question_count": stats["question_count"],
            "active_students": stats["user_count"],
            "active_plans": stats["active_plans"],
            "timezone_name": active_settings.timezone_name,
        }

    @app.get("/api/questions")
    async def list_questions(admin_token: str | None = Header(default=None, alias="X-Admin-Token")) -> dict[str, list[dict[str, Any]]]:
        validate_admin_token(active_settings, admin_token)
        return {"questions": repository.list_questions()}

    @app.post("/api/questions")
    async def create_question(
        image: UploadFile = File(...),
        question_number: str = Form(default=""),
        correct_option: str = Form(...),
        caption: str = Form(default=""),
        topic: str = Form(default="عام"),
        difficulty: str = Form(default="متوسط"),
        admin_token: str | None = Header(default=None, alias="X-Admin-Token"),
    ) -> dict[str, Any]:
        validate_admin_token(active_settings, admin_token)

        parsed_question_number = parse_positive_integer_input(question_number)
        if parsed_question_number is None:
            raise HTTPException(status_code=422, detail="أدخل رقم سؤال صحيحًا أكبر من صفر.")

        normalized_option = correct_option.strip().upper()
        if normalized_option not in VALID_OPTIONS:
            raise HTTPException(status_code=422, detail="الإجابة الصحيحة يجب أن تكون أ أو ب أو ج أو د.")

        if image.content_type is None or not image.content_type.startswith("image/"):
            raise HTTPException(status_code=400, detail="يرجى رفع ملف صورة صالح.")

        image_bytes = await image.read()
        if not image_bytes:
            raise HTTPException(status_code=400, detail="ملف الصورة المرفوع فارغ.")

        return repository.save_question(
            question_number=parsed_question_number,
            caption=caption,
            topic=topic,
            difficulty=difficulty,
            correct_option=normalized_option,
            original_filename=image.filename or "upload.jpg",
            image_bytes=image_bytes,
        )

    @app.post("/api/questions/import-pdf")
    async def import_pdf_questions(
        pdf: UploadFile = File(...),
        answer_sheet: UploadFile | None = File(default=None),
        start_question_number: str = Form(default=""),
        conflict_strategy: str = Form(default=""),
        topic: str = Form(default="عام"),
        difficulty: str = Form(default="متوسط"),
        admin_token: str | None = Header(default=None, alias="X-Admin-Token"),
    ) -> dict[str, Any]:
        validate_admin_token(active_settings, admin_token)

        filename = pdf.filename or "import.pdf"
        content_type = (pdf.content_type or "").lower()
        if content_type not in {"application/pdf", "application/x-pdf"} and not filename.lower().endswith(".pdf"):
            raise HTTPException(status_code=400, detail="يرجى رفع ملف PDF صالح.")

        pdf_temp_path, pdf_size = await save_upload_to_temporary_path(pdf, ".pdf")
        if pdf_size <= 0:
            pdf_temp_path.unlink(missing_ok=True)
            raise HTTPException(status_code=400, detail="ملف PDF المرفوع فارغ.")

        answer_key_by_question_number = None
        starting_question_number = parse_positive_integer_input(start_question_number)
        if starting_question_number is None:
            pdf_temp_path.unlink(missing_ok=True)
            raise HTTPException(status_code=422, detail="أدخل رقم أول سؤال في ملف PDF.")

        normalized_conflict_strategy = conflict_strategy.strip().lower()
        if normalized_conflict_strategy and normalized_conflict_strategy not in VALID_CONFLICT_STRATEGIES:
            pdf_temp_path.unlink(missing_ok=True)
            raise HTTPException(status_code=422, detail="خيار معالجة التعارض غير صالح.")

        has_answer_sheet = answer_sheet is not None and bool((answer_sheet.filename or "").strip())
        answer_sheet_temp_path: Path | None = None

        try:
            if has_answer_sheet:
                answer_sheet_suffix = Path(answer_sheet.filename or "answers.xlsx").suffix or ".xlsx"
                answer_sheet_temp_path, answer_sheet_size = await save_upload_to_temporary_path(answer_sheet, answer_sheet_suffix)
                if answer_sheet_size <= 0:
                    raise HTTPException(status_code=400, detail="ملف Excel المرفوع فارغ.")

                answer_key_by_question_number = extract_answer_key_from_sheet(
                    answer_sheet.filename or "answers.xlsx",
                    answer_sheet_temp_path,
                )

            page_count = read_pdf_page_count(pdf_temp_path)
            conflict_numbers = [starting_question_number + offset for offset in range(page_count)]
            conflicting_questions = repository.list_active_questions_by_numbers(conflict_numbers)
            if conflicting_questions and normalized_conflict_strategy not in VALID_CONFLICT_STRATEGIES:
                raise HTTPException(status_code=409, detail=build_question_number_conflict_detail(conflicting_questions))

            imported_questions, failed_pages, unanswered_question_numbers, ready_count, skipped_existing_question_numbers, replaced_question_numbers = repository.import_pdf_questions(
                original_filename=filename,
                pdf_path=pdf_temp_path,
                starting_question_number=starting_question_number,
                topic=topic,
                difficulty=difficulty,
                answer_key_by_question_number=answer_key_by_question_number,
                existing_question_strategy=normalized_conflict_strategy or "error",
            )
            return {
                "imported_count": len(imported_questions),
                "ready_count": ready_count,
                "draft_question_numbers": unanswered_question_numbers,
                "skipped_existing_question_numbers": skipped_existing_question_numbers,
                "replaced_question_numbers": replaced_question_numbers,
                "failed_pages": failed_pages,
                "questions": imported_questions,
            }
        finally:
            pdf_temp_path.unlink(missing_ok=True)
            if answer_sheet_temp_path is not None:
                answer_sheet_temp_path.unlink(missing_ok=True)

    @app.patch("/api/questions/{question_id}")
    async def update_question(
        question_id: int,
        payload: dict[str, str],
        admin_token: str | None = Header(default=None, alias="X-Admin-Token"),
    ) -> dict[str, Any]:
        validate_admin_token(active_settings, admin_token)
        correct_option = str(payload.get("correct_option", ""))
        updated_question = repository.update_question_correct_option(question_id, correct_option)
        if updated_question is None:
            raise HTTPException(status_code=404, detail="السؤال غير موجود.")
        return updated_question

    @app.delete("/api/questions/{question_id}")
    async def delete_question(
        question_id: int,
        admin_token: str | None = Header(default=None, alias="X-Admin-Token"),
    ) -> dict[str, Any]:
        validate_admin_token(active_settings, admin_token)
        archived = repository.archive_question(question_id)
        if archived is None:
            raise HTTPException(status_code=404, detail="السؤال غير موجود.")
        return archived

    @app.delete("/api/questions")
    async def delete_all_questions(
        admin_token: str | None = Header(default=None, alias="X-Admin-Token"),
    ) -> dict[str, Any]:
        validate_admin_token(active_settings, admin_token)
        return repository.archive_all_questions()

    @app.post("/api/questions/restore")
    async def restore_questions(
        payload: dict[str, list[int]],
        admin_token: str | None = Header(default=None, alias="X-Admin-Token"),
    ) -> dict[str, Any]:
        validate_admin_token(active_settings, admin_token)
        question_ids = payload.get("question_ids", [])
        if not isinstance(question_ids, list):
            raise HTTPException(status_code=422, detail="قائمة الأسئلة المطلوب التراجع عنها غير صالحة.")
        return repository.restore_questions(question_ids)

    return app


def create_app() -> FastAPI:
    return build_application()


def main() -> None:
    settings = load_settings()
    uvicorn.run(
        "bot:create_app",
        host=settings.host,
        port=settings.port,
        factory=True,
        reload=False,
    )


if __name__ == "__main__":
    main()